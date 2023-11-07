
#Check if all Packages are installed

import sys
import subprocess
import pkg_resources

required = {'PySimpleGUI','openpyxl'}
installed = {pkg.key for pkg in pkg_resources.working_set}
missing = required - installed

if missing:
    python = sys.executable
    subprocess.check_call([python, '-m', 'pip', 'install', *missing], stdout=subprocess.DEVNULL)


#After all packages are installed, start importing libraries

from ast import If, Try
import enum
from logging.handlers import TimedRotatingFileHandler
from archicad import ACConnection
from typing import List, Tuple, Iterable
from itertools import cycle
from collections import Counter

import os, uuid

from openpyxl import Workbook, load_workbook

from difflib import SequenceMatcher

import PySimpleGUI as sg

conn = ACConnection.connect()
assert conn


acc = conn.commands
act = conn.types
acu = conn.utilities

scriptFolder = os.path.dirname(os.path.realpath(__file__))

zones = acc.GetElementsByType('Zone')
objects = acc.GetElementsByType("Object")

#print(objects)
#print(objects[0].elementId.guid)

sg.theme("SystemDefault1")

def GetBuiltInPropertyIds (PropertyName, elements):
    zoneNumberBuiltInIds = act.BuiltInPropertyUserId(PropertyName)
    #print(zoneNumberBuiltInIds)
    zoneNumberIds = acc.GetPropertyIds([zoneNumberBuiltInIds])[0].propertyId
    propertyValuesForZones = acc.GetPropertyValuesOfElements(elements, [zoneNumberIds])
    return propertyValuesForZones


def GetUserDefinedPropertyIds (PropertyGroup, PropertyName , elements):
    try:
        UserDefinedIds = acu.GetUserDefinedPropertyId(PropertyGroup, PropertyName)
        #print(zoneNumberUserDefinedIds)
        #Ids = acc.GetPropertyIds([UserDefinedIds])
        #print(Ids)
        propertyValues = acc.GetPropertyValuesOfElements(elements, [UserDefinedIds])
        return propertyValues

    except:
        
        sg.popup_ok("Zone/Object Property called " + PropertyName + " located in folder" + PropertyGroup+", was not found in Archicad. Please make sure this property exists in your Archicad file before running the code again!")
        os._exit()

#define method that extracts specific property from elements
def GetPropertyValues(_propertyValuesForElements):
    propertyValues_ = []
    j = 0
    for l in _propertyValuesForElements:
        if('normal' in _propertyValuesForElements[j].propertyValues[0].propertyValue.status):
            #print(isinstance(_propertyValuesForElements[j].propertyValues[0].propertyValue.value,str))
            if(isinstance(_propertyValuesForElements[j].propertyValues[0].propertyValue.value,str) or isinstance(_propertyValuesForElements[j].propertyValues[0].propertyValue.value,int) or isinstance(_propertyValuesForElements[j].propertyValues[0].propertyValue.value,float)):
              # check if property is SingleEnum type
                 
                 propertyValue = _propertyValuesForElements[j].propertyValues[0].propertyValue.value
                 propertyValues_.append(propertyValue)
            else:
                 propertyValue = _propertyValuesForElements[j].propertyValues[0].propertyValue.value.displayValue
                 propertyValues_.append(propertyValue)
            j = j+1
        else:
            #print(_propertyValuesForElements[j].propertyValues[0].propertyValue.status)
            #print(_propertyValuesForElements[j])
            propertyValues_.append('Error')
            j = j+1
    return propertyValues_


#######Definition for Forderfahige Wohnungen######
def CheckApartmentSize(area,roomNumber,bathtub, buffer):
 
    if(roomNumber  > 0 & roomNumber <=5):
        if((roomNumber-1)*15 + bathtub*5 + 50 + buffer > area): ### this is the general formula for Forderfahigkeit, works with any number of rooms
            return True
        else:
            return False
    else:
        if(roomNumber  > 0):
            print('Room Number is either too low or too high')

def UpdateLoadingState(state, objectToUpdate):
    
    state = state +3
    objectToUpdate.update('█'*state)
    
    return (state)
    
   
'''
try:
    propID = GetUserDefinedPropertyIds("räume", "Wohnungstyp", zones)
except:
    sg.popup_ok("Zone property called 'Wohnungstyp', located in folder 'Räume', was not found in Archicad. Please make sure this property exists in your Archicad file before running the code again!")
    os._exit()
'''





####################################LoadingScreen##############################

layout = [[sg.Text('', size=(60, 1), relief='sunken', font=('Courier', 11),
    key='TEXT')]]
window = sg.Window('Loading...', layout, finalize=True)
text = window['TEXT']
state = 0
while True:
    #print("started")
    
    event, values = window.read(timeout = 100)

    #if event == sg.WINDOW_CLOSED:
        #break

        ######################################################### Read Zone and Object Properties ###################################################

    #print("got to zones")
    zoneNumberPropertyValues = GetBuiltInPropertyIds ('Zone_ZoneNumber', zones)
    #print("read zones")
    #UpdateLoadingState(state)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)
     
    
    
    zoneNamePropertyValues = GetBuiltInPropertyIds ('Zone_ZoneName', zones)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)
     

    zoneLayerNamePropertyValues = GetBuiltInPropertyIds ('ModelView_LayerName', zones)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)
     

    zoneAreaPropertyValues = GetBuiltInPropertyIds ('Zone_CalculatedArea', zones)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)
     

    try:
        zoneHomeStoryNumberPropertyValues = GetUserDefinedPropertyIds ('Python','Home Story Number', zones)
        state = UpdateLoadingState(state, text)
        event, values = window.read(timeout = 100)
    except:
        
        sg.popup_ok("Zone property called 'Home Story Number', located in folder 'Python', was not found in Archicad. Please make sure this property exists in your Archicad file before running the code again!")
        os._exit()

     
    
     
    try:
        zoneWohnungsartPropertyValues = GetUserDefinedPropertyIds ('Räume','Wohnungsart', zones)
        state = UpdateLoadingState(state, text)
        event, values = window.read(timeout = 100)
    except:
        
        sg.popup_ok("Zone property called 'Wohnungsart', located in folder 'Räume', was not found in Archicad. Please make sure this property exists in your Archicad file before running the code again!")
        os._exit()

    try:
        zoneWohnungstypPropertyValues = GetUserDefinedPropertyIds ("Räume", "Wohnungstyp", zones)
    except:
        sg.popup_ok("Zone property called 'Wohnungstyp', located in folder 'Räume', was not found in Archicad. Please make sure this property exists in your Archicad file before running the code again!")
        os._exit()
        
     
    try:
        zoneBuildingNumberPropertyValues = GetUserDefinedPropertyIds ('Allgemeine Werte','Gebäude', zones)
        state = UpdateLoadingState(state, text)
        event, values = window.read(timeout = 100)
    except:
        sg.popup_ok("Zone property called 'Gebäude', located in folder 'Allgemeine Werte', was not found in Archicad. Please make sure this property exists in your Archicad file before running the code again!")
        os._exit()


    #Get needed properties
    zoneNumbers = GetPropertyValues (zoneNumberPropertyValues)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)
     

    zoneNames = GetPropertyValues (zoneNamePropertyValues)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)
         

    zoneLayers = GetPropertyValues (zoneLayerNamePropertyValues)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)
         

    zoneAreas = GetPropertyValues (zoneAreaPropertyValues)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)
         

    zoneHomeStoryNumbers = GetPropertyValues (zoneHomeStoryNumberPropertyValues)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)
         

    zoneWohnungsart = GetPropertyValues (zoneWohnungsartPropertyValues)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)
         

    zoneBuidingNumber = GetPropertyValues (zoneBuildingNumberPropertyValues)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)

    objectNamePropertyValues = GetBuiltInPropertyIds ('IdAndCategories_Name', objects)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)

    objectNames = GetPropertyValues(objectNamePropertyValues)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)

    
         


    objectLayerPropertyValues = GetBuiltInPropertyIds ('ModelView_LayerName', objects)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)
         

    try:
        objectHomeStoryNumberPropertyValues = GetUserDefinedPropertyIds ('Python','Home Story Number', objects)
        state = UpdateLoadingState(state, text)
        event, values = window.read(timeout = 100)
        #print('This is the slow property')
        
    except:
        sg.popup_ok("Object property called 'Home Story Number', located in 'Python' folder, was not found in Archicad. Please make sure this property exists in your Archicad file before running the code again!")
        os._exit()
         

    
         

    
         



    objectLayers = GetPropertyValues(objectLayerPropertyValues)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)
         

    objectHomeStoryNumbers = GetPropertyValues(objectHomeStoryNumberPropertyValues)
    state = UpdateLoadingState(state, text)
     
    event, values = window.read(timeout = 100)
         

    
     
    
         

    
         

    

    window.close()
    break


        ######################################################### Read Zone Properties ###################################################
    

    


####################################LoadingScreen##############################










######MainWindow#######

layerIds = acc.GetAttributesByType('Layer')
layerAttributes = acc. GetLayerAttributes(layerIds)

allLayers = []


for i, currentLayer in enumerate (layerAttributes):
    currentLayer = layerAttributes[i].layerAttribute.name
    allLayers.append(currentLayer)
allLayers.sort()




zoneNamesWindow = []#zoneNames
zoneLayersWindow = zoneLayers

#zoneNamesWindow = list(set(zoneNamesWindow))
zoneLayersWindow = list(set(zoneLayersWindow))

#zoneNamesWindow.sort()
zoneLayersWindow.sort()

objectLayersWindow = objectLayers
objectNamesWindow = []#objectNames

objectLayersWindow = list(set(objectLayersWindow))
#objectNamesWindow = list(set(objectNamesWindow))

objectLayersWindow.sort()
#objectNamesWindow.sort()



col_1 = [
    [sg.CB("Check Zone Names?", key= "-ZoneNames-")],
    [sg.CB("Check Amount of Rooms?", key= "-RoomAmount-")],
    [sg.CB("Check Förderfähigkeit?", key= "-Förderfähigkeit-")],
    
    #[sg.CB("Check Excel?", key= "-RoomAmount-", disabled = True)]
    ]

col_2 = [
    [sg.CB("Overwrite?", key= "-OverwriteZoneNames-", disabled = True)],
    [sg.CB("Overwrite?", key= "-OverwriteRoomAmount-", disabled = True)],
    [sg.CB("Overwrite?", key= "-OverwriteFörderfähigkeit-", disabled = True)],
    #[sg.CB("Overwrite?", key= "-OverwriteExcel-", disabled = True)]
    ]

col_3 = [
    [sg.T('Zones - Layers:')],
    [sg.Listbox(values= zoneLayersWindow, size = (30,10), select_mode= sg.LISTBOX_SELECT_MODE_MULTIPLE, key = '-ListboxZoneLayers-', disabled = True)],
    [sg.Button('Add', key= '-Add-ZoneLayers-', disabled = True), sg.Button('Reset', key= '-Reset-ZoneLayers-', disabled = True)],
    [sg.T('Accepted room names in an apartment:')],
    [sg.Listbox(values= zoneNamesWindow, size=(30,10),select_mode= sg.LISTBOX_SELECT_MODE_MULTIPLE, key = '-ListboxZoneNames-', disabled = True)],
    [sg.Button('Add', key= '-Add-ZoneNames-', disabled = True),sg.Button('Reset', key= '-Reset-ZoneNames-', disabled = True)],
    [sg.T('Possible Layers for Bathtub :')],
    [sg.Listbox(values = objectLayersWindow, size = (30,10), select_mode= sg.LISTBOX_SELECT_MODE_MULTIPLE, key = '-ListboxObjectLayers-', disabled = True)],
    [sg.Button('Add', key= '-Add-ObjectLayers-', disabled = True), sg.Button('Reset', key= '-Reset-ObjectLayers-', disabled = True)],
    [sg.T('Possible Name for Bathtub :')],
    [sg.Listbox(values= objectNamesWindow, size = (30,10), select_mode= sg.LISTBOX_SELECT_MODE_MULTIPLE, key = '-ListboxObjectNames-', disabled = True)],
    [sg.Button('Add', key= '-Add-ObjectNames-', disabled = True), sg.Button('Reset', key= '-Reset-ObjectNames-', disabled = True)],
    #[sg.T('Excel Files:')],
    #[sg.Listbox(values= "", size = (30,10), select_mode= sg.LISTBOX_SELECT_MODE_MULTIPLE, disabled = True)],
    #[sg.FileBrowse(button_text = "Browse", disabled = True, file_types= (("Excel Files", "*.xlsx"), ("ALL Files", "*.*")))]
    ]

col_4 = [
    [sg.T('Selected Zones - Layers:')],
    [sg.Listbox(size=(30,10), select_mode= sg.LISTBOX_SELECT_MODE_MULTIPLE, key = '-Selected ZoneLayers-', values = [])],
    [sg.Button('Delete', key = '-DeleteSelectedZoneLayers-')],
    [sg.T('Selected Zone - Names:')],
    [sg.Listbox(size=(30,10), select_mode= sg.LISTBOX_SELECT_MODE_MULTIPLE, key = '-Selected ZoneNames-', values = [])],
    [sg.Button('Delete', key = '-DeleteSelectedZoneNames-')],
    [sg.T('Selected Bathtub Layers:')],
    [sg.Listbox(size=(30,10), select_mode= sg.LISTBOX_SELECT_MODE_MULTIPLE, key = '-Selected ObjectLayers-', values = [])],
    [sg.Button('Delete', key = '-DeleteSelectedObjectLayers-')],
    [sg.T('Selected Bathtub Names:')],
    [sg.Listbox(size=(30,10), select_mode= sg.LISTBOX_SELECT_MODE_MULTIPLE, key = '-Selected ObjectNames-', values = [])],
    [sg.Button('Delete', key = '-DeleteSelectedObjectNames-')],
    #[sg.T('Selected Excel - Files:')],
    #[sg.Multiline(size=(30,10))],
    #[sg.Button('Delete', key = '-DeleteExcel-')]
    ]

layout = [
    [sg.Col(col_1), sg.Col(col_2), sg.VerticalSeparator(), sg.Col(col_3),sg.Col(col_4)],
    [sg.Ok(), sg.Cancel()]
    ]

updatedZoneLayers = []
updatedZoneNames = []
updatedObjectLayers = []
updatedObjectNames = []
window = sg.Window("Settings", layout)

CheckRoomAmount = False
CheckZoneNames = False
CheckEligibility = False
OverwriteRoomAmount = False
OverwriteEligibility = False
OverwriteZoneNames = False




RunProgram = False

while True:
    event, values = window.read(timeout = 100)
    
    if values["-RoomAmount-"] == True:

        CheckRoomAmount = True
        
        window["-OverwriteRoomAmount-"].update(disabled = False)


        window['-ListboxZoneLayers-'].update(disabled = False)
        window['-Add-ZoneLayers-'].update(disabled = False)
        window['-Selected ZoneLayers-'].update(disabled = False)
        window['-Reset-ZoneLayers-'].update(disabled = False)
        window['-Selected ZoneLayers-'].update(disabled = False)
        window['-DeleteSelectedZoneLayers-'].update(disabled = False)
        

        window['-ListboxZoneNames-'].update(disabled = False)
        window['-Add-ZoneNames-'].update(disabled = False)
        window['-Selected ZoneNames-'].update(disabled = False)
        window['-Reset-ZoneNames-'].update(disabled = False)
        window['-Selected ZoneNames-'].update(disabled = False)
        window['-DeleteSelectedZoneNames-'].update(disabled = False)
        
        

        window['-ListboxObjectLayers-'].update(disabled = False)
        window['-Add-ObjectLayers-'].update(disabled = False)
        window['-Selected ObjectLayers-'].update(disabled = False)
        window['-Reset-ObjectLayers-'].update(disabled = False)
        window['-Selected ObjectLayers-'].update(disabled = False)
        window['-DeleteSelectedObjectLayers-'].update(disabled = False)

        window['-ListboxObjectNames-'].update(disabled = False)
        window['-Add-ObjectNames-'].update(disabled = False)
        window['-Selected ObjectNames-'].update(disabled = False)
        window['-Reset-ObjectNames-'].update(disabled = False)
        window['-Selected ObjectNames-'].update(disabled = False)
        window['-DeleteSelectedObjectNames-'].update(disabled = False)

    if values["-RoomAmount-"] == False:

        CheckRoomAmount = False
        window["-OverwriteRoomAmount-"].update(False)
        window["-OverwriteRoomAmount-"].update(disabled = True)

        if values["-Förderfähigkeit-"] == False:

            updatedObjectLayers = []
            updatedObjectNames = []
            updatedZoneLayers = []
            updatedZoneNames = []

            
            window['-Selected ObjectLayers-'].update("")
            window['-Selected ObjectNames-'].update("")
            window['-Selected ZoneLayers-'].update("")
            window['-Selected ZoneNames-'].update("")

            window['-ListboxZoneLayers-'].update(disabled = True)
            window['-Add-ZoneLayers-'].update(disabled = True)
            window['-Selected ZoneLayers-'].update(disabled = True)
            window['-Reset-ZoneLayers-'].update(disabled = True)
            window['-Selected ZoneLayers-'].update(disabled = True)
            window['-DeleteSelectedZoneLayers-'].update(disabled = True)

            window['-ListboxZoneNames-'].update(disabled = True)
            window['-Add-ZoneNames-'].update(disabled = True)
            window['-Selected ZoneNames-'].update(disabled = True)
            window['-Reset-ZoneNames-'].update(disabled = True)
            window['-Selected ZoneNames-'].update(disabled = True)
            window['-DeleteSelectedZoneNames-'].update(disabled = True)

            window['-ListboxObjectLayers-'].update(disabled = True)
            window['-Add-ObjectLayers-'].update(disabled = True)
            window['-Selected ObjectLayers-'].update(disabled = True)
            window['-Reset-ObjectLayers-'].update(disabled = True)
            window['-Selected ObjectLayers-'].update(disabled = True)
            window['-DeleteSelectedObjectLayers-'].update(disabled = True)

            

            window['-ListboxObjectNames-'].update(disabled = True)
            window['-Add-ObjectNames-'].update(disabled = True)
            window['-Selected ObjectNames-'].update(disabled = True)
            window['-Reset-ObjectNames-'].update(disabled = True)
            window['-Selected ObjectNames-'].update(disabled = True)
            window['-DeleteSelectedObjectNames-'].update(disabled = True)

            

    if values["-Förderfähigkeit-"] == True:

        CheckEligibility = True

        window["-OverwriteFörderfähigkeit-"].update(disabled = False)


        window['-ListboxZoneLayers-'].update(disabled = False)
        window['-Add-ZoneLayers-'].update(disabled = False)
        window['-Selected ZoneLayers-'].update(disabled = False)
        window['-Reset-ZoneLayers-'].update(disabled = False)
        window['-Selected ZoneLayers-'].update(disabled = False)
        window['-DeleteSelectedZoneLayers-'].update(disabled = False)
        

        window['-ListboxZoneNames-'].update(disabled = False)
        window['-Add-ZoneNames-'].update(disabled = False)
        window['-Selected ZoneNames-'].update(disabled = False)
        window['-Reset-ZoneNames-'].update(disabled = False)
        window['-Selected ZoneNames-'].update(disabled = False)
        window['-DeleteSelectedZoneNames-'].update(disabled = False)
        
        

        window['-ListboxObjectLayers-'].update(disabled = False)
        window['-Add-ObjectLayers-'].update(disabled = False)
        window['-Selected ObjectLayers-'].update(disabled = False)
        window['-Reset-ObjectLayers-'].update(disabled = False)
        window['-Selected ObjectLayers-'].update(disabled = False)
        window['-DeleteSelectedObjectLayers-'].update(disabled = False)

        window['-ListboxObjectNames-'].update(disabled = False)
        window['-Add-ObjectNames-'].update(disabled = False)
        window['-Selected ObjectNames-'].update(disabled = False)
        window['-Reset-ObjectNames-'].update(disabled = False)
        window['-Selected ObjectNames-'].update(disabled = False)
        window['-DeleteSelectedObjectNames-'].update(disabled = False)

    if values["-Förderfähigkeit-"] == False:

        window["-OverwriteFörderfähigkeit-"].update(False)
        window["-OverwriteFörderfähigkeit-"].update(disabled = True)

        CheckEligibility = False



    




    if values["-OverwriteRoomAmount-"] == True:

        OverwriteRoomAmount = True

    if values["-OverwriteRoomAmount-"] == False:

        OverwriteRoomAmount = False

    if values["-OverwriteFörderfähigkeit-"] == True:

        OverwriteEligibility = True

    if values["-OverwriteFörderfähigkeit-"] == False:

        OverwriteEligibility = False

    if values["-OverwriteZoneNames-"] == True:

        OverwriteZoneNames = True

    if values["-OverwriteZoneNames-"] == False:

        OverwriteZoneNames = False

    if values["-ZoneNames-"] == True:

        CheckZoneNames = True

    if values["-ZoneNames-"] == False:

        CheckZoneNames = False

    

        



        

    if event == '-ListboxZoneLayers-':
        selection = values[event]

    if event == '-Add-ZoneLayers-':
        existingZonesLayers = updatedZoneLayers
        #print(existingZonesLayers)
        newLayers = values['-ListboxZoneLayers-']
        updatedZoneLayers = list(set(existingZonesLayers).union(set(newLayers)))
        updatedZoneLayers.sort()
        window['-Selected ZoneLayers-'].update(updatedZoneLayers)
        window['-ListboxZoneLayers-'].update(zoneLayersWindow)
        #print(updatedZoneLayers)

        zoneNamesWindow = []
        for i, name in enumerate(zoneNames):
            if zoneLayers[i] in updatedZoneLayers:
                zoneNamesWindow.append(zoneNames[i])

        zoneNamesWindow = list(set(zoneNamesWindow))
        zoneNamesWindow.sort()
        window['-ListboxZoneNames-'].update(zoneNamesWindow) # having some problems with updating the lists dynamically - maybe there is a problem with a tuple?
        
    if event == '-Reset-ZoneLayers-':
        window['-ListboxZoneLayers-'].update(zoneLayersWindow)
        window['-Selected ZoneLayers-'].update('')

    if event == '-DeleteSelectedZoneLayers-':
        selected = window['-Selected ZoneLayers-'].get()
        updatedZoneLayers = list(set(updatedZoneLayers) - set(selected))
        updatedZoneLayers.sort()
        window['-Selected ZoneLayers-'].update(updatedZoneLayers)
        #print(updatedZoneLayers)

        zoneNamesWindow = []
        
        for i, name in enumerate(zoneNames):
            if zoneLayers[i] in updatedZoneLayers:
                zoneNamesWindow.append(zoneNames[i])

        zoneNamesWindow = list(set(zoneNamesWindow))
        zoneNamesWindow.sort()
        window['-ListboxZoneNames-'].update(zoneNamesWindow)





    if event == '-ListboxZoneNames-':
        selection = values[event]

    if event == '-Add-ZoneNames-':
        existingZonesNames = updatedZoneNames
        #print(existingZonesNames)
        newZoneNames = values['-ListboxZoneNames-']
        updatedZoneNames = list(set(existingZonesNames).union( set(newZoneNames)))
        updatedZoneNames.sort()
        window['-Selected ZoneNames-'].update(updatedZoneNames)
        window['-ListboxZoneNames-'].update(zoneNamesWindow)
        #print(updatedZoneNames)

    if event == '-Reset-ZoneNames-':
        window['-ListboxZoneNames-'].update(zoneNamesWindow)

    if event == '-DeleteSelectedZoneNames-':
        selected = window['-Selected ZoneNames-'].get()
        updatedZoneNames = list(set(updatedZoneNames) - set(selected))
        updatedZoneNames.sort()
        window["-Selected ZoneNames-"].update(updatedZoneNames)
        #print(updatedZoneNames)




    if event == '-ListboxObjectLayers-':
        selection = values[event]

    if event == '-Add-ObjectLayers-':
        existingObjectLayers = updatedObjectLayers
        #print(existingObjectLayers)
        newObjectLayers = values['-ListboxObjectLayers-']
        updatedObjectLayers = list(set(existingObjectLayers).union( set(newObjectLayers)))
        updatedObjectLayers.sort()
        window['-Selected ObjectLayers-'].update(updatedObjectLayers)
        window['-ListboxObjectLayers-'].update(objectLayersWindow)
        #print(updatedObjectLayers)

        objectNamesWindow = []
        
        for i, name in enumerate(objectNames):
            if objectLayers[i] in updatedObjectLayers:
                objectNamesWindow.append(objectNames[i])

        objectNamesWindow = list(set(objectNamesWindow))
        objectNamesWindow.sort()
        window['-ListboxObjectNames-'].update(objectNamesWindow)

    if event == '-Reset-ObjectLayers-':
        window['-ListboxObjectLayers-'].update(objectLayersWindow)

    if event == '-DeleteSelectedObjectLayers-':
        selected = window['-Selected ObjectLayers-'].get()
        updatedObjectLayers = list(set(updatedObjectLayers) - set(selected))
        updatedObjectLayers.sort()
        window['-Selected ObjectLayers-'].update(updatedObjectLayers)
        #print(updatedObjectLayers)

        objectNamesWindow = []
        
        for i, name in enumerate(objectNames):
            if objectLayers[i] in updatedObjectLayers:
                objectNamesWindow.append(objectNames[i])

        objectNamesWindow = list(set(objectNamesWindow))
        objectNamesWindow.sort()
        window['-ListboxObjectNames-'].update(objectNamesWindow)




    if event == '-ListboxObjectNames-':
        selection = values[event]

    if event == '-Add-ObjectNames-':
        existingObjectNames = updatedObjectNames
        #print(existingObjectNames)
        newObjectNames = values['-ListboxObjectNames-']
        updatedObjectNames = list(set(existingObjectNames).union( set(newObjectNames)))
        updatedObjectNames.sort()
        window['-Selected ObjectNames-'].update(updatedObjectNames)
        window['-ListboxObjectNames-'].update(objectNamesWindow)
        #print(updatedObjectNames)

    if event == '-Reset-ObjectNames-':
        window['-ListboxObjectNames-'].update(objectNamesWindow)

    if event == '-DeleteSelectedObjectNames-':
        selected = window['-Selected ObjectNames-'].get()
        updatedObjectNames = list(set(updatedObjectNames) - set(selected))
        updatedObjectNames.sort()
        window['-Selected ObjectNames-'].update(updatedObjectNames)
        #print(updatedObjectNames)

    


    if event == 'Ok':
        RunProgram = True
        if(CheckZoneNames or CheckRoomAmount or CheckEligibility):
            if ((len(updatedZoneLayers)) == 0):
                sg.popup_ok("No Layers selected for Zones")
            else:
                if((len(updatedZoneNames)) == 0 and (CheckRoomAmount or CheckEligibility)):
                    sg.popup_ok("No Names selected for Zones")
                else:
                    if((len(updatedObjectLayers)) == 0 and CheckEligibility):
                        sg.popup_ok("No Layers selected for Bathtub")
                    else:
                        if((len(updatedObjectNames)) == 0 and CheckEligibility):
                            sg.popup_ok("No Names selected for Bathtub")
                        else:
                            window.Close()
                            break
        else:
            sg.popup_ok("No functions selected, program unable to run")

    if event == 'Cancel':
        window.Close()
        break

    if event == sg.WIN_CLOSED:
        window.Close()
        quit()
        break
window.Close()

possibleRoomNames = (updatedZoneNames)
bathtub_PossibleLayerNames = (updatedObjectLayers)
bathtub_PossibleObjectNames = (updatedObjectNames)
zone_PossibleLayerNames = (updatedZoneLayers)

'''
print(possibleRoomNames)
print(bathtub_PossibleLayerNames)
print(bathtub_PossibleObjectNames)
print(zone_PossibleLayerNames)

print(len(bathtub_PossibleObjectNames))
'''

#################Filter and read only properties related to bathtubs - and save time with it ##############

objects_Bathtubs = []
bathtub_HomeStoryNumbers = []
bathtub_ObjectNames = []
bathtub_ObjectLayers = []




for i, name in enumerate(objectNames):
    for possibleName in bathtub_PossibleObjectNames:
        if(name == possibleName):
            objects_Bathtubs.append(objects[i])
            bathtub_HomeStoryNumbers.append( objectHomeStoryNumbers[i])
            bathtub_ObjectNames.append(objectNames[i])
            bathtub_ObjectLayers.append(objectLayers[i])
            break
            
#for _object in objects_Bathtubs:


bathtubRelatedToZoneNumberPropertyValues = GetBuiltInPropertyIds ('IdAndCategories_RelatedZoneNumber', objects_Bathtubs)
#state = UpdateLoadingState(state, text)
     
#event, values = window.read(timeout = 100)
         

bathtubRelatedToZoneNamePropertyValues = GetBuiltInPropertyIds ('IdAndCategories_RelatedZoneName', objects_Bathtubs)
#state = UpdateLoadingState(state, text)
     
#event, values = window.read(timeout = 100)

bathtubRelatedToZoneNumbers = GetPropertyValues(bathtubRelatedToZoneNumberPropertyValues)
#state = UpdateLoadingState(state, text)
     
#event, values = window.read(timeout = 100)
         

bathtubRelatedToZoneNames = GetPropertyValues(bathtubRelatedToZoneNamePropertyValues)
#state = UpdateLoadingState(state, text)

#event, values = window.read(timeout = 100)





        


############################################### CONFIGURATION #####################################################
if RunProgram:
    outputFolder = scriptFolder
    excelFileName = "230201_403_Berechnung Wohnfläche Nutzfläche Stellplätze.xlsx"




    bufferForApartmentArea = 0.5

    overwriteProperties = False
    overwriteExcel = False  

    #possibleRoomNames = ['Wohnen','Zimmer', 'Schlafen']

    #bathtub_PossibleLayerNames = ["50 Möblierung", "50 Sanitäreinrichtung"]

    #zone_PossibleLayerNames = ["70 Raum"]

    #bathtub_PossibleObjectNames = ['Badewanne',]

    #list with possible apartment starting numbers to be used later (1.1.1, 3.4.1 etc)
    possibleApartmentStartingNumbers = [1,2,3,4,5,6]


    ############################################# CONFIGURATION ################################################################


    #propertyValueStringPrefix = ''

   

    #Couple of UsefulPropertyNames:
    #'General_RelatedZoneNumber'
    #'General_RelatedZoneName'
    #IdAndCategories_RelatedZoneName
    #'IdAndCategories_RelatedZoneNumber'
    #'IdAndCategories_Name'
    #'General_ToZoneNumber'
    #'General_ToZone'
    #'General_FromZoneNumber'
    #'General_FromZone'


    ########################################## Filter Zones ############################################




    #filter out Zones that are not part of Apartments
    apartmentZones = []  
    apartmentZoneNumbers = []
    apartmentZoneNames = []
    apartmentZoneAreas = []
    apartmentZoneWohnungsart = []
    apartmentBuildingNumber = []


    for i, currentZone in enumerate(zones):


        currentLayerName = zoneLayers[i]
        currentHomeStoryNumber = int(zoneHomeStoryNumbers[i])
        currentZoneNumber = zoneNumbers[i]
        currentZoneName = zoneNames[i]
        currentZoneArea = zoneAreas[i]
        currentZoneWohnungsart = zoneWohnungsart[i]
        currentBuildingNumber = zoneBuidingNumber[i]


        #check HomeStory
        if(currentHomeStoryNumber >= 0):

            #Check if first character in ZoneNumber is among PossibleStartingNumbers
            if len(currentZoneNumber)>0 :
                for possibleStartingNumber in possibleApartmentStartingNumbers:
                    firstCharacter = currentZoneNumber[0]
                    if str(possibleStartingNumber) in firstCharacter:

                        #Check if layer name for current zone is among PossibleLayerNames
                            for possibleLayerName in zone_PossibleLayerNames:
                                if(currentLayerName == possibleLayerName):
                                    apartmentZones.append(currentZone)
                                    apartmentZoneNumbers.append(currentZoneNumber)
                                    apartmentZoneNames.append(currentZoneName)
                                    apartmentZoneAreas.append(currentZoneArea)
                                    apartmentZoneWohnungsart.append(currentZoneWohnungsart)
                                    apartmentBuildingNumber.append(currentBuildingNumber)

  





    ###################################### Read Objects, their Properties, then Filter ########################################
    



    testLength = len(bathtub_ObjectLayers)

    if(testLength != len(bathtub_HomeStoryNumbers)):
        print('Error with length of objectHomeStoryNumbers')
    if(testLength != len(bathtubRelatedToZoneNumbers)):
        print('Error with length of bathtubRelatedToZoneNumbers')
    if(testLength != len(bathtubRelatedToZoneNames)):
        print('Error with length of bathtubRelatedToZoneNames')
    if(testLength != len(bathtub_ObjectNames)):
        print('Error with length of objectNames')



    bathtubs = []
    bathtub_RelatedToZoneNumbers = []
    bathtub_RelatedToZoneNames = []
    bathtub_IncorrectZoneNames = []

    for i, currentObject in enumerate(objects_Bathtubs):
    
        #objectLayerPropertyValue = GetBuiltInPropertyIds ('ModelView_LayerName', objects[i])
        currentLayerName = bathtub_ObjectLayers[i]
        currentObjectName =bathtub_ObjectNames[i]
        #if('normal' in objectHomeStoryNumberPropertyValues[i].propertyValues[0].propertyValue.status):
            #for some reason some objects cannot access Home Story Number in AC, we need to skip these
        if('Error' not in str(bathtub_HomeStoryNumbers[i])):
            currentHomeStoryNumber = int(bathtub_HomeStoryNumbers[i])
            for possibleLayerName in bathtub_PossibleLayerNames:
                if(currentLayerName == possibleLayerName):
                    if(currentHomeStoryNumber >= 0):
                        for k, currentPossibleName in enumerate(bathtub_PossibleObjectNames): #"Badewanne"
                            if(currentPossibleName in currentObjectName):
                                bathtubs.append(currentObject)
                                bathtub_RelatedToZoneNumbers.append(bathtubRelatedToZoneNumbers[i])
                                bathtub_RelatedToZoneNames.append(bathtubRelatedToZoneNames[i])
                                if(CheckZoneNames):
                                    #if('Wannenbad' not in bathtubRelatedToZoneNamePropertyValues[i].propertyValues[0].propertyValue.value ):
                                    if('Wannenbad' not in bathtubRelatedToZoneNames[i] ):
                                        print('Following Bathtub is not in a correctly named Zone - it should be called Wannenbad:')
                                        if(bathtubRelatedToZoneNumbers[i] == ''):
                                                print('No ZoneNumber has been found related to Bathtub')
                                                
                                        else:
                                                print(bathtubRelatedToZoneNumbers[i])
                                        print('Current name is: ' + str(bathtubRelatedToZoneNames[i]))
                                        bathtub_IncorrectZoneNames += [bathtubRelatedToZoneNumbers[i], bathtubRelatedToZoneNames[i]]

        else:
            for i, currentPossibleName in enumerate(bathtub_PossibleObjectNames): #"Badewanne'
                if(currentPossibleName in currentObjectName):
                    print("Object reading error")
                    print("Following object has been included, even though Home Story could not be read :") # the fact that home story cannot be read has to do with Hotlinks. Reading all Bathtub objects without this might not be a problem. 
                    print(bathtub_ObjectNames[i])
                    print(bathtubRelatedToZoneNumbers[i])
                    print(bathtub_HomeStoryNumbers[i])
                    print('')
                    bathtubs.append(currentObject)
                    bathtub_RelatedToZoneNumbers.append(bathtubRelatedToZoneNumbers[i])
                    bathtub_RelatedToZoneNames.append(bathtubRelatedToZoneNames[i])
                    if(CheckZoneNames):
                        if('Wannenbad' not in bathtubRelatedToZoneNamePropertyValues[i].propertyValues[0].propertyValue.value ):
                            print('Following Bathtub is not in a correctly named Zone - it should be called Wannenbad:')
                            if(bathtubRelatedToZoneNumbers[i] == ''):
                                    print('No ZoneNumber has been found related to Bathtub')
                            else:
                                    print(bathtubRelatedToZoneNumbers[i])
                            print('Current name is: ' + str(bathtubRelatedToZoneNames[i]))
                            bathtub_IncorrectZoneNames += [bathtubRelatedToZoneNumbers[i], bathtubRelatedToZoneNames[i]]

    #check similarity between zone names
    if(CheckZoneNames):
        def removeDigits(a):
            #result = ''.join([i for i in a if not i.isdigit()])
        
            result = ''.join([i for i in a if i.isalpha()])
            result.replace(" ","")
        
            return result

        def similar(a, b):
            return SequenceMatcher(None, removeDigits(a), removeDigits(b)).ratio()

    

        misspelledZones = []
        zoneNamesSet= list(set(zoneNames))
        zoneNamesSet.sort()

        for  name1 in zoneNamesSet:
            for name2 in zoneNamesSet:
                if((similar (name1, name2)> 0.75) & (similar (name1, name2) < 1.00)):
                   print('Following room names might be misspelled:' + name1+ ' and '+name2)
                   print(removeDigits(name1)+' '+ removeDigits(name2))
                   print(similar (name1, name2))
                   misspelledZones.append( "'" + str(name1+ "' and '"+name2 + ".' Similarity - " +str(round(similar (name1, name2), 2)) ))

        col_1 = [
        [sg.T("These rooms might be misspelled")],
        [sg.T("Please check and correct if necessary, before proceeding.")],
        [sg.T("Incorrectly named rooms might lead to errors in further calculations")],
        [sg.Listbox(values = misspelledZones, size = (60,20))],
        ]

        col_2 = [
        [sg.T("These Bathrooms might be incorrectly named")],
        [sg.Listbox(values = bathtub_IncorrectZoneNames, size = (60,20))],
        ]

        layout = [
        [sg.Col(col_1)],
        [ sg.Col(col_2)],
        [sg.Ok(), sg.Cancel()]
        ]
        window = sg.Window("RoomNumbers", layout)

        while True:
            event, values = window.read(timeout= 100)
            if event == "Ok":
                
                window.Close()

                break
            if event == "Cancel":
                window.Close()
                quit()
            if event == sg.WIN_CLOSED:
                window.Close()
                quit()
                break
               



    ###################################### Start Data Processing #######################################

    #create list of bools containing whether a zone is a "room" or not
    apartmentRoomBools = []


    #check whether zones are "rooms"

    def CheckIfZonesAreRooms(zoneNameToBeChecked, possibleRoomNames):
        finalResult = 0
        for possibleRoomName in possibleRoomNames:
            if(possibleRoomName in zoneNameToBeChecked):
                finalResult = 1
     
        return finalResult
    """
    #k=0
    #for j, zoneName in enumerate(zoneNames):
    #while k < len(zoneNames):
        # apartmentRoomBools.append(0)
       # l = 0
       # while l <  len(possibleRoomNames):
            #if(possibleRoomNames[l] in zoneNames[k]):
        #apartmentRoomBools.append( CheckIfZonesAreRooms(zoneName, possibleRoomNames))

            #l+=1
        #k+=1
    """
    for  zoneName in apartmentZoneNames:
        apartmentRoomBools.append(CheckIfZonesAreRooms(zoneName, possibleRoomNames))




    #create list with unique Zone Numbers
    allUniqueApartmentNumbers = list(set(apartmentZoneNumbers))
    allUniqueApartmentNumbers.sort()


    #Create empty dictionary with Apartment Numbers and Quantities of Rooms (for now only 0)

    apartmentRoomAmountDict = dict.fromkeys(allUniqueApartmentNumbers, 0 )

    #Create empty dictionary with Apartment Numbers and Quantities of Bathtubs (for now only 0)

    apartmentBathtubAmountDict = dict.fromkeys(allUniqueApartmentNumbers, 0 )

    #Create empty dictionary with Apartment Numbers and Area (for now only 0)

    apartmentSizeDict = dict.fromkeys(allUniqueApartmentNumbers, 0 )

    #Create empty dictionary with Apartment Numbers and Area (for now only 0)

    apartmentWohnungsartDict = dict.fromkeys(allUniqueApartmentNumbers, '')

    #Create empty dictionary with Apartment Numbers and Area (for now only 0)

    apartmentBuildingNumberDict = dict.fromkeys(allUniqueApartmentNumbers, '')



    #check and sum room quantities for apartments
    for currentNumber in allUniqueApartmentNumbers:
        l=0
        while l < len (apartmentRoomBools):
            if currentNumber == apartmentZoneNumbers[l]:
                currentQuantityOfRooms = apartmentRoomAmountDict[currentNumber]
                apartmentRoomAmountDict[currentNumber] = currentQuantityOfRooms +  apartmentRoomBools[l]
        
            l+=1

    #find groups without any "Rooms"
    #print('groups without any "Rooms"')

    apartmentsWithZeroRooms = []
    #k=0
    if(CheckRoomAmount):
        for currentKey  in list(apartmentRoomAmountDict.keys()):
            #for currentValue in list(apartmentRoomAmountDict.values()):
            
                #myvalues = list( apartmentRoomAmountDict.values())
               # mykeys = list(apartmentRoomAmountDict.keys())
            currentValue = apartmentRoomAmountDict[currentKey]
            if(currentValue == 0):
                apartmentsWithZeroRooms.append(currentKey)
                #print ("This Apartment has 0 rooms - " + currentKey)

    
    zonesWithZeroRooms = []

    for currentApartmentNumber in apartmentsWithZeroRooms:
        for i, currentZoneNumber in enumerate(apartmentZoneNumbers):
            if currentApartmentNumber == currentZoneNumber:
                zonesWithZeroRooms.append([currentApartmentNumber,apartmentZoneNames[i]])



        

            
            #k+=1

    #check if apartment has bathtub
    for currentApartmentNumber in allUniqueApartmentNumbers:
        for currentObjectNumber in bathtub_RelatedToZoneNumbers:
            if currentObjectNumber == currentApartmentNumber:
                apartmentBathtubAmountDict[currentApartmentNumber] = 1


    #calculate apartment area
    for currentApartmentNumber in allUniqueApartmentNumbers:
        for i, currentZoneNumber in enumerate (apartmentZoneNumbers):
                if currentZoneNumber == currentApartmentNumber:
                    currentSizeOfApartment = apartmentSizeDict[currentZoneNumber]
                    apartmentSizeDict[currentZoneNumber] = currentSizeOfApartment + apartmentZoneAreas[i]
    
    propID = acu.GetUserDefinedPropertyId("räume", "Wohnungstyp")
    propertyValuesForZones = acc.GetPropertyValuesOfElements(apartmentZones, [propID])

    detailedValue = acc.GetDetailsOfProperties([propID])


    enumValues = []
    i=0
    j=1
    for element in enumerate(detailedValue[0].propertyDefinition.possibleEnumValues): #potential for improvement
        if('Zimmer' in detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.displayValue):
            digit = ''.join([k for k in detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.displayValue if k.isdigit()])
            if(digit == str(j)):
                #if(str(j) in detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.displayValue):
                    enumValues.append(detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.enumValueId)
                    j+=1
        i+=1
    #print(enumValues)



    #################################### Start Data Exporting #######################################
    A= []
    B= []
    C= []
    D= []
    apartmentsWithWrongRooms = () #(apartment number, apartment room amount, apartment property in archicad)
    print('')
    if(CheckRoomAmount):
        if OverwriteRoomAmount:
            #i=0
            elemPropertyValues = []
            for i, currentZone in enumerate(zones):
                    j=0
                    for j, currentKey in enumerate(list (apartmentRoomAmountDict.keys())):
                        if list(apartmentRoomAmountDict.values())[j] !=0:
                            if zoneNumbers[i] == currentKey:
                                x = list (apartmentRoomAmountDict.values())[j] - 1
                                propValue = act.NormalSingleEnumPropertyValue(enumValues[x]) # this needs to be checked, if property input list in Archicad changes
                                elemPropertyValues.append(act.ElementPropertyValue(
                                    currentZone.elementId, propID, propValue))
                                #A += [str(list (apartmentRoomAmountDict.keys())[j])]
                                #B += [str(list (apartmentRoomAmountDict.values())[j])]
                                #C += [str((propertyValuesForZones[i].propertyValues[0].propertyValue.value.displayValue))] #cant figure out this triple tuple...
                                #D += [str(apartmentZoneNames[i])]
                        #j+=1
                    #i+=1
            
        #else: 
        i=0
        print('Following Apartments need to be updated in Archicad with Quantity of Rooms: ')
        for zone in apartmentZones:
                j=0
                while j < len(list (apartmentRoomAmountDict.keys())):
                    if apartmentZoneNumbers[i] == list (apartmentRoomAmountDict.keys())[j]:
                        x = list (apartmentRoomAmountDict.values())[j] - 1 
                        if (list (apartmentRoomAmountDict.values())[j] != 0): # if room quantity is 0 then dont show
                            if ((propertyValuesForZones[i].propertyValues[0].propertyValue.value.displayValue) != enumValues[x].displayValue): # the property in Archicad is not the same as counted by the algorithm
                                print(str(list (apartmentRoomAmountDict.keys())[j]) + ' - '+str ( list (apartmentRoomAmountDict.values())[j]) + ' Zimmer')
                                print("At the moment it is "+ propertyValuesForZones[i].propertyValues[0].propertyValue.value.displayValue)
                                print('')
                                #apartmentsWithWrongRooms +=  (str(list (apartmentRoomAmountDict.keys())[j]),),+tuple((str(list (apartmentRoomAmountDict.values())[j]),str((propertyValuesForZones[i].propertyValues[0].propertyValue.value.displayValue))))
                                A += [str(list (apartmentRoomAmountDict.keys())[j])]
                                B += [str(list (apartmentRoomAmountDict.values())[j])]
                                C += [str((propertyValuesForZones[i].propertyValues[0].propertyValue.value.displayValue))] #cant figure out this triple tuple...
                                D += [str(apartmentZoneNames[i])]
                        
                    j+=1
                i+=1
        print('')


        B = [x for _, x in sorted(zip(A,B))]
        C = [x for _, x in sorted(zip(A,C))]
        D = [x for _, x in sorted(zip(A,D))]
        A.sort()
        apartmentsWithWrongRooms = zip(A,B,C,D)

        text = []
        previous_a = ""
        for a,b in zonesWithZeroRooms:
                if a != previous_a:
                    text.append(a)
                    text.append("      " + b)
                else:
                    text.append("      " + b)
                previous_a = a

        col_1 = [
        [sg.T("These Apartments have 0 rooms:")],
        [sg.Listbox(values = text, size = (60,20))],
        ]

        text = []
        previous_a = ""
        for a,b,c, d in apartmentsWithWrongRooms:
            if(a !=previous_a):
                text.append(" Apartment number '" + a + "' has " + b + " rooms.")
                text.append("        '"+ d + "' in Archicad is set as " + c)
            else:
                text.append("        '"+ d + "' in Archicad is set as " + c)
            previous_a = a
        col_2 = [
            [sg.T("These Apartments need to be updated/ will be overwritten in Archicad:")],
            [sg.Listbox(values = text, size = (60,20))],
        ]
        layout = [
            [sg.Col(col_1)],
            [sg.Col(col_2)],
            [sg.Ok(), sg.Cancel()]
            ]
        window = sg.Window("RoomNumbers", layout)

        while True:
            event, values = window.read(timeout = 100)
            if event == "Ok":
                if (OverwriteRoomAmount and len(apartmentsWithWrongRooms)>0) :
                    proceed= sg.popup_yes_no("Are you sure you want to overwrite these zone properties? Press no if you want to skip overwrite", title = 'Continue?')
                    if(proceed == "Yes"):
                        window.Close()
                        break
                    if(proceed == "No"):
                        OverwriteRoomAmount = False
                        window.Close()
                        break
                else:
                    window.Close()
                    break
            if event == "Cancel":
                window.Close()
                quit()
            if event == sg.WIN_CLOSED:
                window.Close()
                quit()
                


    if OverwriteRoomAmount:
        acc.SetPropertyValuesOfElements(elemPropertyValues)
    





    propID = acu.GetUserDefinedPropertyId("räume", "Wohnungsart")
    propertyValuesForZones = acc.GetPropertyValuesOfElements(zones, [propID])

    detailedValue = acc.GetDetailsOfProperties([propID])


    wohnungsartPropList = ['gefördert', 'freifinanziert', 'Eigentum']
    wohnungsartPropDict = dict.fromkeys(wohnungsartPropList, 0)

    for i, element in enumerate (detailedValue[0].propertyDefinition.possibleEnumValues):
        if('förd' in detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.displayValue):
             wohnungsartPropDict["gefördert"] = detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.enumValueId
        if('freifinanziert' in detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.displayValue):
             wohnungsartPropDict["freifinanziert"] = detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.enumValueId
        if('Eigentum' in detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.displayValue):
             wohnungsartPropDict["Eigentum"] = detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.enumValueId
    

    #ask if property should be overwritten - done
    #overwrite Wohnungsart property - done
    #print list if it should not be overwritten - done

    #organize Wohnungsart into Dictionary
    for currentApartmentNumber in allUniqueApartmentNumbers:
        for i, currentZoneNumber in enumerate(apartmentZoneNumbers):
            if(currentApartmentNumber == currentZoneNumber):
                if CheckApartmentSize(apartmentSizeDict[currentApartmentNumber], apartmentRoomAmountDict[currentApartmentNumber], apartmentBathtubAmountDict [currentApartmentNumber], bufferForApartmentArea): #if förderfähig then
                    apartmentWohnungsartDict[currentApartmentNumber] = [True, '']
                    break
                else:
                    apartmentWohnungsartDict[currentApartmentNumber] = [False, '']
    
    #apartmentSameWohnungsart = []
    currentApartmentWohnungsart = []
    for currentApartmentNumber in allUniqueApartmentNumbers:
        #keys = []
        #values = []
        #sortedKeys = []
        #sortedValues = []
        currentApartmentWohnungsart = []
        for i, currentZoneNumber in enumerate(apartmentZoneNumbers):
            if(currentApartmentNumber == currentZoneNumber):
                currentApartmentWohnungsart.append(apartmentZoneWohnungsart[i])

        keys = Counter(currentApartmentWohnungsart).keys()
        values = Counter(currentApartmentWohnungsart).values()
        sortedKeys = [x for _, x in reversed(sorted(zip(values,keys)))]
        sortedValues =  [x for x, _ in reversed(sorted(zip(values,keys)))]

        if(len((Counter(currentApartmentWohnungsart)).keys()) > 1):
           
            if(sortedValues[0]>sortedValues[1]):
                apartmentWohnungsartDict[currentApartmentNumber][1]=sortedKeys[0]
            else:
                apartmentWohnungsartDict[currentApartmentNumber][1]= "Unclear"
        else:
            apartmentWohnungsartDict[currentApartmentNumber][1]=sortedKeys[0]


    [x for _, x in sorted(zip(A,B))]


    #check if apartment if foerderfahig                     #check this part, there might be errors!
    
    elemPropertyValues = []

    correctEligibleZones = []
    incorrectEligibleZones = []
    correctIneligibleZones = []
    incorrectIneligibleZones = []

    for currentApartmentNumber in allUniqueApartmentNumbers:
        
        if(apartmentWohnungsartDict[currentApartmentNumber][0]): #if förderfähig then

            print()
            print('Following apartment förderfähig: ' + currentApartmentNumber)
            print(str(apartmentRoomAmountDict[currentApartmentNumber]) + ' Zimmer - ' + str(round(apartmentSizeDict[currentApartmentNumber],2)) + ' m3')
            
            allZonesCorrect = True
            allZonesSame = True
            currentWohnungsart = []
            currentApartmentRooms = []

            #counterCorrect = 0
            #counterIncorrect = 0

            if(apartmentRoomAmountDict[currentApartmentNumber] != 0):

                for i, currentZone in enumerate(apartmentZones):

                            if(apartmentZoneNumbers[i] == currentApartmentNumber):
                                
                                if len(currentWohnungsart) == 0:
                                    currentWohnungsart = apartmentZoneWohnungsart[i]
                                else:
                                    #if(not allZonesSame):
                                    if(apartmentZoneWohnungsart[i] != currentWohnungsart):
                                        allZonesSame = False
                                        allZonesCorrect = False

                                if("förd" not in apartmentZoneWohnungsart[i]): # if zone is not set as gefördert
                                    if(("Eigentum" not in apartmentZoneWohnungsart[i]) and ("freifinanziert" not in apartmentZoneWohnungsart[i])): #smth else than the three
                                        print ( 'Property needs to be corrected in Archicad! Check zone number: ' + apartmentZoneNumbers[i] + " - " + apartmentZoneNames[i] + ". At the moment it is set as " + str(apartmentZoneWohnungsart[i]))
                                        if OverwriteEligibility:
                                            
                                            propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict["gefördert"])
                                            elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                        allZonesCorrect = False

                                        #incorrectEligibleZones += ["    "+str(currentApartmentNumber + " - " + apartmentZoneNames[i] + " - " + apartmentZoneWohnungsart[i])]
                                        #counterIncorrect += 1
                                        currentApartmentRooms += ["    "+str(currentApartmentNumber + " - " + apartmentZoneNames[i] + " - " + apartmentZoneWohnungsart[i])]
                                    else:
                                        if OverwriteEligibility:
                                            if ("Eigentum" in apartmentWohnungsartDict[currentApartmentNumber][1]):
                                                propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict["Eigentum"])
                                                elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                            else:
                                                if("freifinanziert" in apartmentWohnungsartDict[currentApartmentNumber][1]):
                                                    propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict["freifinanziert"])
                                                    elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                        #if(("Eigentum" in apartmentZoneWohnungsart[i]) or ("freifinanziert" in apartmentZoneWohnungsart[i])):
                                        #incorrectEligibleZones += [str(currentApartmentNumber + " - " + apartmentZoneNames[i])]
                                        #counterIncorrect += 1
                                            #correctEligibleZones += [str(currentApartmentNumber + " - " + apartmentZoneNames[i] + " -" + apartmentZoneWohnungsart[i])]
                                            #counterCorrect += 1
                                        currentApartmentRooms +=["    "+str(currentApartmentNumber + " - " + apartmentZoneNames[i]+ " - " + apartmentZoneWohnungsart[i])]
                                        allZonesCorrect = False



                                
                                else:
                                    if OverwriteEligibility:
                                        if ("Eigentum" in apartmentWohnungsartDict[currentApartmentNumber][1]):
                                                    propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict["Eigentum"])
                                                    elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                        else:
                                            if("freifinanziert" in apartmentWohnungsartDict[currentApartmentNumber][1]):
                                                propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict["freifinanziert"])
                                                elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                    #if("förd" in apartmentWohnungsartDict[currentApartmentNumber][1]):
                                        #apartmentWohnungsartDict[currentApartmentNumber][1] = "gefördert"
                                    #correctEligibleZones += [str(currentApartmentNumber + " - " + apartmentZoneNames[i]+ " -" + apartmentZoneWohnungsart[i])]
                                    #counterCorrect += 1
                                    currentApartmentRooms += ["    "+str(currentApartmentNumber + " - " + apartmentZoneNames[i] + " - " + apartmentZoneWohnungsart[i])]
                                    

                                    
                                    #print("Apartment is eligible, and property is correctly set in Archicad")
                if allZonesSame:
                    #if(counterCorrect>0)
                        #del correctEligibleZones[-counterCorrect:]
                    #if(counterIncorrect>0)
                        #del incorrectEligibleZones [-counterIncorrect:]
                    if allZonesCorrect == False:
                        
                        print("Not all Zones are set as 'gefördert'. This might or might not be intentional")
                        incorrectEligibleZones.append(str(currentApartmentNumber + " - Zones are not set as 'gefördert'"))
                        for room in currentApartmentRooms:
                            incorrectEligibleZones.append(room)
                    else:
                    
                        print("All Zones in Archicad are correctly set as 'gefördert'.")
                        correctEligibleZones+= [str(currentApartmentNumber + " - "+apartmentWohnungsartDict[currentApartmentNumber][1])]
                else:
                
                    print("Not all Zones are set with the same property. Please check.")
                    incorrectEligibleZones.append(str(currentApartmentNumber + " - Not all Zones are set with the same property. Please check."))
                    for room in currentApartmentRooms:
                        incorrectEligibleZones.append(room)

                
                

        else:

            currentApartmentRooms = []
            if(apartmentRoomAmountDict[currentApartmentNumber] != 0):
                
                print()
                print('Following apartment NOT förderfähig: ' + currentApartmentNumber)
                print(str(apartmentRoomAmountDict[currentApartmentNumber]) + ' Zimmer - ' + str(round(apartmentSizeDict[currentApartmentNumber],2)) + ' m3')

                #counterCorrect = 0
                #counterIncorrect = 0
                        
                allZonesCorrect = True
                allZonesSame = True
                currentWohnungsart = []

                for i, currentZone in enumerate(apartmentZones):
                    
                    if(apartmentZoneNumbers[i] == currentApartmentNumber):
                        if len(currentWohnungsart) == 0:
                            currentWohnungsart = apartmentZoneWohnungsart[i]
                        else:
                            if(apartmentZoneWohnungsart[i] != currentWohnungsart):
                                allZonesSame = False
                        if("förd" in apartmentZoneWohnungsart[i]): # is zone wrongly set as "gefördert"?
                            if OverwriteEligibility:
                                    if("förd" in apartmentWohnungsartDict[currentApartmentNumber][1] or 'freifinanziert' in apartmentWohnungsartDict[currentApartmentNumber][1]):
                                
                                        propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict['freifinanziert'])
                                        elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                    else:
                                        if('Eigentum' in apartmentWohnungsartDict[currentApartmentNumber][1]):
                                
                                            propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict['Eigentum'])
                                            elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                        else:
                                            if "Unclear" not in apartmentWohnungsartDict[currentApartmentNumber][1]:
                                                propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict['freifinanziert'])
                                                elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )

                            print ('Property needs to be corrected in Archicad! Check zone number: ' + apartmentZoneNumbers[i] + " - " + apartmentZoneNames[i] + ". At the moment it is set as " +str(apartmentZoneWohnungsart[i]))
                            allZonesCorrect = False

                            #counterIncorrect += 1
                            #incorrectIneligibleZones += ["    "+str(currentApartmentNumber + " - " + apartmentZoneNames[i]+ " - " + apartmentZoneWohnungsart[i])]
                            currentApartmentRooms +=["    "+str(currentApartmentNumber + " - " + apartmentZoneNames[i]+ " - " + apartmentZoneWohnungsart[i])]
                        else:
                            if("Eigentum" in apartmentZoneWohnungsart[i]): #if not eigentum then freifinanziert
                                if OverwriteEligibility:
                                    if("förd" in apartmentWohnungsartDict[currentApartmentNumber][1] or 'freifinanziert' in apartmentWohnungsartDict[currentApartmentNumber][1]):
                                
                                        propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict['freifinanziert'])
                                        elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                    else:
                                        if('Eigentum' in apartmentWohnungsartDict[currentApartmentNumber][1]):
                                
                                            propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict['Eigentum'])
                                            elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                        else:
                                            if "Unclear" not in apartmentWohnungsartDict[currentApartmentNumber][1]:
                                                propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict['freifinanziert'])
                                                elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                #apartmentWohnungsartDict[currentApartmentNumber][1] = 'Eigentum'
                                #counterCorrect += 1
                                #correctIneligibleZones += [str(currentApartmentNumber + " - " + apartmentZoneNames[i]+ " -" + apartmentZoneWohnungsart[i])]
                                currentApartmentRooms +=["    "+str(currentApartmentNumber + " - " + apartmentZoneNames[i]+ " - " + apartmentZoneWohnungsart[i])]
                            else:
                                if("freifinanziert" in apartmentZoneWohnungsart[i]): #if not freifinanziert, then smth is wrong
                                    apartmentWohnungsartDict[currentApartmentNumber][1] = 'freifinanziert'
                                    if OverwriteEligibility:
                                        if("förd" in apartmentWohnungsartDict[currentApartmentNumber][1] or 'freifinanziert' in apartmentWohnungsartDict[currentApartmentNumber][1]):
                                
                                            propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict['freifinanziert'])
                                            elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                        else:
                                            if('Eigentum' in apartmentWohnungsartDict[currentApartmentNumber][1]):
                                
                                                propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict['Eigentum'])
                                                elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                            else:
                                                if "Unclear" not in apartmentWohnungsartDict[currentApartmentNumber][1]:
                                                    propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict['freifinanziert'])
                                                    elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                    #counterCorrect += 1
                                    #correctIneligibleZones += [str(currentApartmentNumber + " - " + apartmentZoneNames[i]+ " -" + apartmentZoneWohnungsart[i])]
                                    currentApartmentRooms +=["    "+str(currentApartmentNumber + " - " + apartmentZoneNames[i]+ " - " + apartmentZoneWohnungsart[i])]
                                else:
                                    print('Property needs to be corrected in Archicad! Check zone number: ' + apartmentZoneNumbers[i] + " - " + apartmentZoneNames[i] + ". At the moment it is set as " + str(apartmentZoneWohnungsart[i]))
                                    allZonesCorrect = False
                                    if OverwriteEligibility:
                                
                                        if("förd" in apartmentWohnungsartDict[currentApartmentNumber][1] or 'freifinanziert' in apartmentWohnungsartDict[currentApartmentNumber][1]):
                                
                                            propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict['freifinanziert'])
                                            elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                        else:
                                            if('Eigentum' in apartmentWohnungsartDict[currentApartmentNumber][1]):
                                
                                                propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict['Eigentum'])
                                                elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
                                            else:
                                                if "Unclear" not in apartmentWohnungsartDict[currentApartmentNumber][1]:
                                                    propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict['freifinanziert'])
                                                    elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )

                                    #counterIncorrect += 1
                                    #incorrectIneligibleZones += ["    "+str(currentApartmentNumber + " - " + apartmentZoneNames[i]+ " - " + apartmentZoneWohnungsart[i])]
                                    currentApartmentRooms +=["    "+str(currentApartmentNumber + " - " + apartmentZoneNames[i]+ " - " + apartmentZoneWohnungsart[i])]

                                
                if allZonesSame:
                    #if(counterCorrect>0)
                        #del correctIneligibleZones[-counterCorrect:]
                    #if(counterIncorrect>0)
                        #del incorrectIneligibleZones[-counterIncorrect:]
                    if allZonesCorrect != True:
                        print("Zones are NOT correct, please check")
                        incorrectIneligibleZones.append(str(currentApartmentNumber) + "Apartment NOT förderfähig")
                        for room in currentApartmentRooms:
                            incorrectIneligibleZones.append(room)
                    else:
                        print("All Zones in Archicad are set correctly! - " + apartmentWohnungsartDict[currentApartmentNumber][1])
                        correctIneligibleZones += [str(currentApartmentNumber + " - "+apartmentWohnungsartDict[currentApartmentNumber][1])]
                else:
                    print("Not all Zones are set with the same property. Please check.") ############### This section needs to be checked again
                    incorrectIneligibleZones.append(str(currentApartmentNumber + " - Not all Zones are set with the same property. Please check."))
                    for room in currentApartmentRooms:
                        incorrectIneligibleZones.append( room)

                        

                            
                    """
            else:
                

                    #if(apartmentWohnungsartDict[currentApartmentNumber][1] != 'freifinanziert' and apartmentWohnungsartDict[currentApartmentNumber][1] != 'Eigentum'): #check if property is correctly set in Archicad
                       print('Property needs to be corrected in Archicad!')

                    else:
                        print('Property is correctly set in Archicad')
                    print('Current Property in Archicad -' + str(apartmentWohnungsartDict[currentApartmentNumber][1]))
                    print('')
                    """
                
            

 
    

    col_1 = [
        [sg.T("Eligible Apartments")],
        [sg.T("Correct apartments:")],
        [sg.Listbox(values = correctEligibleZones, size = (60,20))],
    ]

    col_2 = [
            [sg.T("")],
            [sg.T("Zones that might need to be updated:")],
            [sg.Listbox(values = incorrectEligibleZones, size = (60,20))],
            ]
    """
    text = []
    for a,b,c, d in apartmentsWithWrongRooms:
        text.append(" Apartment number '" + a + "' has " + b + " rooms. Room called '"+ d + "' in Archicad is set as " + c + " rooms.")
    """
    col_3 = [
        [sg.T("Not Eligible Apartments")],
        [sg.T("Correct apartments:")],
        [sg.Listbox(values =  correctIneligibleZones, size = (60,20))],
    ]

    col_4 = [
        [sg.T("")],
        [sg.T("Zones that need to be updated / will be overwritten:")],
        [sg.Listbox(values =  incorrectIneligibleZones, size = (60,20))],
    ]

        
    layout = [
        [sg.Col(col_1), sg.Col(col_2)],
        [sg.Col(col_3), sg.Col(col_4)],
        [sg.Ok(), sg.Cancel()]
        ]
    window = sg.Window("RoomNumbers", layout)

    while True:
        event, values = window.read(timeout= 100)
        if event == "Ok":
            if(OverwriteEligibility and (len(incorrectIneligibleZones) + len(incorrectEligibleZones)) > 0):
                    proceed= sg.popup_yes_no("Are you sure you want to overwrite these zone properties? Press no if you want to skip overwrite", title = 'Continue?')
                    if(proceed == "Yes"):
                        window.Close()
                        break
                    if(proceed == "No"):
                        OverwriteEligibility = False
                        window.Close()
                        break
            else:
                window.Close()
                break
        if event == "Cancel":
            window.Close()
            quit()


    if OverwriteEligibility:
        acc.SetPropertyValuesOfElements(elemPropertyValues)





    

    
    


    #count apartmenttypes (number of rooms, foerderfahigkeit)

    for currentZoneNumber in allUniqueApartmentNumbers:
        for i, currentZone in enumerate(apartmentZoneNumbers):
            if currentZoneNumber == currentZone:
                if apartmentBuildingNumberDict[currentZoneNumber] == '':
                    apartmentBuildingNumberDict[currentZoneNumber] = apartmentBuildingNumber[i]
    
    if overwriteExcel:

        totalAmount_1RoomApt = 0
        totalAmount_2RoomApt = 0
        totalAmount_3RoomApt = 0
        totalAmount_4RoomApt = 0
        totalAmount_5RoomApt = 0

        totalAmount_gefoerdert = 0
        totalAmount_freifinanziert = 0
        totalAmount_Eigentum = 0



        totalAmount_gefoerdert_G1 = 0
        totalAmount_freifinanziert_G1 = 0
        totalAmount_Eigentum_G1 = 0

        totalAmount_gefoerdert_G2 = 0
        totalAmount_freifinanziert_G2 = 0
        totalAmount_Eigentum_G2 = 0

        totalAmount_gefoerdert_G3 = 0
        totalAmount_freifinanziert_G3 = 0
        totalAmount_Eigentum_G3 = 0

        totalAmount_gefoerdert_G4 = 0
        totalAmount_freifinanziert_G4 = 0
        totalAmount_Eigentum_G4 = 0

        totalAmount_gefoerdert_G5 = 0
        totalAmount_freifinanziert_G5 = 0
        totalAmount_Eigentum_G5 = 0

        totalAmount_gefoerdert_G6 = 0
        totalAmount_freifinanziert_G6 = 0
        totalAmount_Eigentum_G6 = 0




        totalArea_gefoerdert = 0
        totalArea_freifinanziert = 0
        totalArea_Eigentum = 0

        totalArea_gefoerdert_G1 = 0
        totalArea_freifinanziert_G1 = 0
        totalArea_Eigentum_G1 = 0

        totalArea_gefoerdert_G2 = 0
        totalArea_freifinanziert_G2 = 0
        totalArea_Eigentum_G2 = 0

        totalArea_gefoerdert_G3 = 0
        totalArea_freifinanziert_G3 = 0
        totalArea_Eigentum_G3 = 0

        totalArea_gefoerdert_G4 = 0
        totalArea_freifinanziert_G4 = 0
        totalArea_Eigentum_G4 = 0

        totalArea_gefoerdert_G5 = 0
        totalArea_freifinanziert_G5 = 0
        totalArea_Eigentum_G5 = 0

        totalArea_gefoerdert_G6 = 0
        totalArea_freifinanziert_G6 = 0
        totalArea_Eigentum_G6 = 0

        for currentApartmentNumber in allUniqueApartmentNumbers:
            if(apartmentRoomAmountDict[currentApartmentNumber] == 1):
                totalAmount_1RoomApt += 1
            if(apartmentRoomAmountDict[currentApartmentNumber] == 2):
                totalAmount_2RoomApt += 1
            if(apartmentRoomAmountDict[currentApartmentNumber] == 3):
                totalAmount_3RoomApt += 1
            if(apartmentRoomAmountDict[currentApartmentNumber] == 4):
                totalAmount_4RoomApt += 1
            if(apartmentRoomAmountDict[currentApartmentNumber] == 5):
                totalAmount_5RoomApt += 1

            if("gefördert" in apartmentWohnungsartDict[currentApartmentNumber]):
                totalAmount_gefoerdert += 1
                totalArea_gefoerdert += apartmentSizeDict[currentApartmentNumber]

                if('1' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    #print('G1 - gefördert')
                    totalAmount_gefoerdert_G1 += 1
                    totalArea_gefoerdert_G1 += apartmentSizeDict[currentApartmentNumber]

                if('2' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    #print('G2 - gefördert')
                    totalAmount_gefoerdert_G2 += 1
                    totalArea_gefoerdert_G2 += apartmentSizeDict[currentApartmentNumber]

                if('3' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_gefoerdert_G3 += 1
                    totalArea_gefoerdert_G3 += apartmentSizeDict[currentApartmentNumber]

                if('4' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_gefoerdert_G4 += 1
                    totalArea_gefoerdert_G4 += apartmentSizeDict[currentApartmentNumber]

                if('5' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_gefoerdert_G5 += 1
                    totalArea_gefoerdert_G5 += apartmentSizeDict[currentApartmentNumber]

                if('6' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_gefoerdert_G6 += 1
                    totalArea_gefoerdert_G6 += apartmentSizeDict[currentApartmentNumber]


            if('freifinanziert' in apartmentWohnungsartDict[currentApartmentNumber]):
                totalAmount_freifinanziert += 1
                totalArea_freifinanziert += apartmentSizeDict[currentApartmentNumber]

                if('1' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_freifinanziert_G1 += 1
                    totalArea_freifinanziert_G1 += apartmentSizeDict[currentApartmentNumber]

                if('2' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_freifinanziert_G2 += 1
                    totalArea_freifinanziert_G2 += apartmentSizeDict[currentApartmentNumber]

                if('3' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_freifinanziert_G3 += 1
                    totalArea_freifinanziert_G3 += apartmentSizeDict[currentApartmentNumber]

                if('4' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_freifinanziert_G4 += 1
                    totalArea_freifinanziert_G4 += apartmentSizeDict[currentApartmentNumber]

                if('5' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_freifinanziert_G5 += 1
                    totalArea_freifinanziert_G5 += apartmentSizeDict[currentApartmentNumber]

                if('6' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_freifinanziert_G6 += 1
                    totalArea_freifinanziert_G6 += apartmentSizeDict[currentApartmentNumber]

            if('Eigentum' in apartmentWohnungsartDict[currentApartmentNumber]):
                totalAmount_Eigentum += 1
                totalArea_Eigentum += apartmentSizeDict[currentApartmentNumber]

                if('1' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_Eigentum_G1 += 1
                    totalArea_Eigentum_G1 += apartmentSizeDict[currentApartmentNumber]
                if('2' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_Eigentum_G2 += 1
                    totalArea_Eigentum_G2 += apartmentSizeDict[currentApartmentNumber]
                if('3' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_Eigentum_G3 += 1
                    totalArea_Eigentum_G3 += apartmentSizeDict[currentApartmentNumber]
                if('4' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_Eigentum_G4 += 1
                    totalArea_Eigentum_G4 += apartmentSizeDict[currentApartmentNumber]
                if('5' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_Eigentum_G5 += 1
                    totalArea_Eigentum_G5 += apartmentSizeDict[currentApartmentNumber]
                if('6' in apartmentBuildingNumberDict[currentApartmentNumber]):
                    totalAmount_Eigentum_G6 += 1
                    totalArea_Eigentum_G6 += apartmentSizeDict[currentApartmentNumber]


    #print(apartmentBuildingNumberDict)
    #print(apartmentWohnungsartDict)

    #export quantities to excel

    
        wb = load_workbook(filename = excelFileName)

        ws = wb["Stellplätze Fahrräder"]

        #C3 - G6 - Eigentum
        ws['C3'] = totalAmount_Eigentum_G6

        #C18 - G1-5 - gefördert + freifinanziert
        ws['C18'] = totalAmount_gefoerdert + totalAmount_freifinanziert # assuming there are none in G6



        ws2 = wb["Stellplätze PKW"]


        # C3 - G6 - Eigentum
        ws2['C3'] = totalAmount_Eigentum_G6

        # C11 - G1-5 - gefördert
        ws2['C11'] = totalAmount_gefoerdert

        # C18 - G1-5 - freifinanziert
        ws2['C18'] = totalAmount_freifinanziert

        ws3 = wb["Wohnflächen"]

        #Anzahl
        #D4 - G1 - gefördert
        ws3['D4'] = totalAmount_gefoerdert_G1
        #D5 - G1 - freifinanziert
        ws3['D5'] = totalAmount_freifinanziert_G1
        #D6 - G1 - Eigentum
        ws3['D6'] = totalAmount_Eigentum_G1
 
        #D11 - G2 - gefördert
        ws3['D11'] = totalAmount_gefoerdert_G2
        #D12 - G2 - freifinanziert
        ws3['D12'] = totalAmount_freifinanziert_G2
        #D13 - G2 - Eigentum
        ws3['D13'] = totalAmount_Eigentum_G2
 
        #D18 - G3 - gefördert
        ws3['D18'] = totalAmount_gefoerdert_G3
        #D19 - G3 - freifinanziert
        ws3['D19'] = totalAmount_freifinanziert_G3
        #D20 - G3 - Eigentum
        ws3['D20'] = totalAmount_Eigentum_G3
 
        #D25 - G4 - gefördert
        ws3['D25'] = totalAmount_gefoerdert_G4
        #D26 - G4 - freifinanziert
        ws3['D26'] = totalAmount_freifinanziert_G4
        #D27 - G4 - Eigentum
        ws3['D27'] = totalAmount_Eigentum_G4
 
        #D32 - G5 - gefördert
        ws3['D32'] = totalAmount_gefoerdert_G5
        #D33 - G5 - freifinanziert
        ws3['D33'] = totalAmount_freifinanziert_G5
        #D34 - G5 - Eigentum
        ws3['D34'] = totalAmount_Eigentum_G5
 
        #D39 - G6 - gefördert
        ws3['D39'] = totalAmount_gefoerdert_G6
        #D40 - G6 - freifinanziert
        ws3['D40'] = totalAmount_freifinanziert_G6
        #D41 - G6 - Eigentum
        ws3['D41'] = totalAmount_Eigentum_G6

        #Wohnfläche
        #C4 - G1 - gefördert
        ws3['C4'] = totalArea_gefoerdert_G1
        #C5 - G1 - freifinanziert
        ws3['C5'] = totalArea_freifinanziert_G1
        #C6 - G1 - Eigentum
        ws3['C6'] = totalArea_Eigentum_G1
 
        #C11 - G2 - gefördert
        ws3['C11'] = totalArea_gefoerdert_G2
        #C12 - G2 - freifinanziert
        ws3['C12'] = totalArea_freifinanziert_G2
        #C13 - G2 - Eigentum
        ws3['C13'] = totalArea_Eigentum_G2
 
        #C18 - G3 - gefördert
        ws3['C18'] = totalArea_gefoerdert_G3
        #C19 - G3 - freifinanziert
        ws3['C19'] = totalArea_freifinanziert_G3
        #C20 - G3 - Eigentum
        ws3['C20'] = totalArea_Eigentum_G3
 
        #C25 - G4 - gefördert
        ws3['C25'] = totalArea_gefoerdert_G4
        #C26 - G4 - freifinanziert
        ws3['C26'] = totalArea_freifinanziert_G4
        #C27 - G4 - Eigentum
        ws3['C27'] = totalArea_Eigentum_G4
 
        #C32 - G5 - gefördert
        ws3['C32'] = totalArea_gefoerdert_G5
        #C33 - G5 - freifinanziert
        ws3['C33'] = totalArea_freifinanziert_G5
        #C34 - G5 - Eigentum
        ws3['C34'] = totalArea_Eigentum_G5
 
        #C39 - G6 - gefördert
        ws3['C39'] = totalArea_gefoerdert_G6
        #C40 - G6 - freifinanziert
        ws3['C40'] = totalArea_freifinanziert_G6
        #C41 - G6 - Eigentum
        ws3['C41'] = totalArea_Eigentum_G6


        outputPath = os.path.join(scriptFolder, excelFileName)
        wb.save(outputPath)
