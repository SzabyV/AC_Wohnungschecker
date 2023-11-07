
import enum
from logging.handlers import TimedRotatingFileHandler
from archicad import ACConnection
from typing import List, Tuple, Iterable
from itertools import cycle

import os, sys, uuid

from openpyxl import Workbook, load_workbook

import PySimpleGUI as sg

conn = ACConnection.connect()
assert conn

acc = conn.commands
act = conn.types
acu = conn.utilities

scriptFolder = os.path.dirname(os.path.realpath(__file__))

######Window#######

layerIds = acc.GetAttributesByType('Layer')
layerAttributes = acc. GetLayerAttributes(layerIds)

allLayers = []


for i, currentLayer in enumerate (layerAttributes):
    currentLayer = layerAttributes[i].layerAttribute.name
    allLayers.append(currentLayer)
allLayers.sort()

sg.theme("DarkBlue13")

col_1 = [
    [sg.CB("Check Amount of Rooms?", key= "-RoomAmount-")],
    [sg.CB("Check Förderfähigkeit?", key= "-Förderfähigkeit-")],
    [sg.CB("Check Zone Names?", key= "-ZoneNames-")],
    [sg.CB("Check Excel?", key= "-RoomAmount-")]
    ]

col_2 = [
    [sg.CB("Overwrite?", key= "-OverwriteRoomAmount-")],
    [sg.CB("Overwrite?", key= "-OverwriteFörderfähigkeit-")],
    [sg.CB("Overwrite?", key= "-OverwriteZoneNames-")],
    [sg.CB("Overwrite?", key= "-OverwriteRoomAmount-")]
    ]

col_3= [
    [sg.T('Zones - Layers:')],
    [sg.Listbox(values= allLayers, size = (30,10), select_mode= sg.LISTBOX_SELECT_MODE_MULTIPLE, key = '-ListboxZoneLayers-')],
    [sg.Button('Add', key= '-Add-ZonesLayers-'), sg.Button('Reset', key= '-Reset-ZonesLayers-')],
    [sg.T('Accepted room names:')],
    [sg.Multiline(size=(30,10))],
    [sg.Button('Reset', key= '-Reset-ZoneNames-')],
    [sg.T('Objects - Layers:')],
    [sg.Listbox(values= allLayers, size = (30,10), select_mode= sg.LISTBOX_SELECT_MODE_MULTIPLE, key = '-ListboxObjectsLayers-')],
    [sg.Button('Add', key= '-Add-ObjectsLayers-'), sg.Button('Reset', key= '-Reset-ZoneLayers-')],
    [sg.T('Excel Files:')],
    [sg.Listbox(values= "", size = (30,10), select_mode= sg.LISTBOX_SELECT_MODE_MULTIPLE)],
    [sg.FileBrowse(button_text = "Browse", file_types= (("Excel Files", "*.xlsx"), ("ALL Files", "*.*")))]
    ]

col_4 = [
    [sg.T('Selected Zones - Layers:')],
    [sg.Listbox(size=(30,10), key = '-Selected Zones-', values = [])],
    [sg.Button('Delete', key = '-DeleteSelectedZone-')],
    [sg.T('Selected Room - Names:')],
    [sg.Multiline(size=(30,10))],
    [sg.Button('Delete')],
    [sg.T('Selected Object - Layers:')],
    [sg.Listbox(size=(30,10), values = [])],
    [sg.Button('Delete')],
    [sg.T('Selected Excel - Files:')],
    [sg.Multiline(size=(30,10))],
    [sg.Button('Delete')]
    ]

layout = [
    [sg.Col(col_1), sg.Col(col_2), sg.VerticalSeparator(), sg.Col(col_3),sg.Col(col_4)],
    [sg.Ok(), sg.Cancel()]
    ]

updatedLayers = []
window = sg.Window("Settings", layout)
while True:
    event, values = window.read()
    if event == '-Add-ObjectsLayers-':
        print("You pressed'-Add-ObjectsLayers-'")
    if event == '-ListboxZoneLayers-':
        selection = values[event]

    if event == '-Add-ZonesLayers-':
        existingLayers = updatedLayers
        print(existingLayers)
        newLayers = values['-ListboxZoneLayers-']
        updatedLayers = list(set(existingLayers).union( set(newLayers)))
        #newtext = ''
        #for layer in layers:
         #   newtext = newtext + layer + '\n'
        updatedLayers.sort()
        window['-Selected Zones-'].update(updatedLayers)
        window['-ListboxZoneLayers-'].update(allLayers)

    if event == '-Reset-ZonesLayers-':
        window['-ListboxZoneLayers-'].update(allLayers)

    if event == '-DeleteSelectedZone-':
        selected = window['-Selected Zones-'].get()
        updatedLayers = list(set(updatedLayers) - set(selected))
        updatedLayers.sort()
        window['-Selected Zones-'].update(updatedLayers)

    if event == sg.WIN_CLOSED:
        break

window.close()
        


############################################### CONFIGURATION #####################################################

outputFolder = scriptFolder
excelFileName = "230201_403_Berechnung Wohnfläche Nutzfläche Stellplätze.xlsx"


zones = acc.GetElementsByType('Zone')
objects = acc.GetElementsByType("Object")

bufferForApartmentArea = 0.5

overwriteProperties = False
overwriteExcel = False  

possibleRoomNames = ['Wohnen','Zimmer', 'Schlafen']

bathtub_PossibleLayerNames = ["50 Möblierung", "50 Sanitäreinrichtung"]

zone_PossibleLayerNames = ["70 Raum"]

#list with possible apartment starting numbers to be used later (1.1.1, 3.4.1 etc)
possibleApartmentStartingNumbers = [1,2,3,4,5,6]


############################################# CONFIGURATION ################################################################


#propertyValueStringPrefix = ''

def GetBuiltInPropertyIds (PropertyName, elements):
    zoneNumberBuiltInIds = act.BuiltInPropertyUserId(PropertyName)
    #print(zoneNumberBuiltInIds)
    zoneNumberIds = acc.GetPropertyIds([zoneNumberBuiltInIds])[0].propertyId
    propertyValuesForZones = acc.GetPropertyValuesOfElements(elements, [zoneNumberIds])
    return propertyValuesForZones


def GetUserDefinedPropertyIds (PropertyGroup, PropertyName , elements):
    UserDefinedIds = acu.GetUserDefinedPropertyId(PropertyGroup, PropertyName)
    #print(zoneNumberUserDefinedIds)
    #Ids = acc.GetPropertyIds([UserDefinedIds])
    #print(Ids)
    propertyValues = acc.GetPropertyValuesOfElements(elements, [UserDefinedIds])
    return propertyValues

#define method that extracts specific property from elements
def GetPropertyValues(_propertyValuesForElements):
    propertyValues_ = []
    j = 0
    for l in _propertyValuesForElements:
        if('normal' in _propertyValuesForElements[j].propertyValues[0].propertyValue.status):
            #print(isinstance(_propertyValuesForElements[j].propertyValues[0].propertyValue.value,str))
            if(isinstance(_propertyValuesForElements[j].propertyValues[0].propertyValue.value,str) or isinstance(_propertyValuesForElements[j].propertyValues[0].propertyValue.value,int) or isinstance(_propertyValuesForElements[j].propertyValues[0].propertyValue.value,float)):
              # check if property is SingleEnum type
                 
                 propertyValue = _propertyValuesForElements[j].propertyValues[0].propertyValue.value
                 propertyValues_.append(propertyValue)
            else:
                 propertyValue = _propertyValuesForElements[j].propertyValues[0].propertyValue.value.displayValue
                 propertyValues_.append(propertyValue)
            j = j+1
        else:
            #print(_propertyValuesForElements[j].propertyValues[0].propertyValue.status)
            #print(_propertyValuesForElements[j])
            propertyValues_.append('Error')
            j = j+1
    return propertyValues_


#######Definition for Forderfahige Wohnungen######
def CheckApartmentSize(area,roomNumber,bathtub, buffer):
 
    if(roomNumber  > 0 & roomNumber <=5):
        if((roomNumber-1)*15 + bathtub*5 + 50 + buffer > area): ### this is the general formula for Forderfahigkeit, works with any number of rooms
            return True
        else:
            return False
    else:
        print('Room Number is either too low or too high')

    

#Couple of UsefulPropertyNames:
#'General_RelatedZoneNumber'
#'General_RelatedZoneName'
#IdAndCategories_RelatedZoneName
#'IdAndCategories_RelatedZoneNumber'
#'IdAndCategories_Name'
#'General_ToZoneNumber'
#'General_ToZone'
#'General_FromZoneNumber'
#'General_FromZone'


##########################################   Read Zones, their Properties, then Filter ############################################


zoneNumberPropertyValues = GetBuiltInPropertyIds ('Zone_ZoneNumber', zones)
zoneNamePropertyValues = GetBuiltInPropertyIds ('Zone_ZoneName', zones)
zoneLayerNamePropertyValues = GetBuiltInPropertyIds ('ModelView_LayerName', zones)
zoneAreaPropertyValues = GetBuiltInPropertyIds ('Zone_CalculatedArea', zones)
zoneHomeStoryNumberPropertyValues = GetUserDefinedPropertyIds ('Python','Home Story Number', zones)
zoneWohnungsartPropertyValues = GetUserDefinedPropertyIds ('räume','Wohnungsart', zones)
zoneBuildingNumberPropertyValues = GetUserDefinedPropertyIds ('Allgemeine Werte','Gebäude', zones)

#Get needed properties
zoneNumbers = GetPropertyValues (zoneNumberPropertyValues)
zoneNames = GetPropertyValues (zoneNamePropertyValues)
zoneLayers = GetPropertyValues (zoneLayerNamePropertyValues)
zoneAreas = GetPropertyValues (zoneAreaPropertyValues)
#print('Error with Zones - Home Story Number:')
zoneHomeStoryNumbers = GetPropertyValues (zoneHomeStoryNumberPropertyValues)
#print('Error with Zones - Wohnungsart:')
zoneWohnungsart = GetPropertyValues (zoneWohnungsartPropertyValues)
#print('Error with Zones - Building Number:')
zoneBuidingNumber = GetPropertyValues (zoneBuildingNumberPropertyValues)


#filter out Zones that are not part of Apartments
apartmentZones = []  
apartmentZoneNumbers = []
apartmentZoneNames = []
apartmentZoneAreas = []
apartmentZoneWohnungsart = []
apartmentBuildingNumber = []


for i, currentZone in enumerate(zones):


    currentLayerName = zoneLayers[i]
    currentHomeStoryNumber = int(zoneHomeStoryNumbers[i])
    currentZoneNumber = zoneNumbers[i]
    currentZoneName = zoneNames[i]
    currentZoneArea = zoneAreas[i]
    currentZoneWohnungsart = zoneWohnungsart[i]
    currentBuildingNumber = zoneBuidingNumber[i]


    #check HomeStory
    if(currentHomeStoryNumber >= 0):

        #Check if first character in ZoneNumber is among PossibleStartingNumbers
        if len(currentZoneNumber)>0 :
            for possibleStartingNumber in possibleApartmentStartingNumbers:
                firstCharacter = currentZoneNumber[0]
                if str(possibleStartingNumber) in firstCharacter:

                    #Check if layer name for current zone is among PossibleLayerNames
                        for possibleLayerName in zone_PossibleLayerNames:
                            if(currentLayerName == possibleLayerName):
                                apartmentZones.append(currentZone)
                                apartmentZoneNumbers.append(currentZoneNumber)
                                apartmentZoneNames.append(currentZoneName)
                                apartmentZoneAreas.append(currentZoneArea)
                                apartmentZoneWohnungsart.append(currentZoneWohnungsart)
                                apartmentBuildingNumber.append(currentBuildingNumber)

  





###################################### Read Objects, their Properties, then Filter ########################################


objectLayerPropertyValues = GetBuiltInPropertyIds ('ModelView_LayerName', objects)
objectHomeStoryNumberPropertyValues = GetUserDefinedPropertyIds ('Python','Home Story Number', objects)
objectRelatedToZoneNumberPropertyValues = GetBuiltInPropertyIds ('IdAndCategories_RelatedZoneNumber', objects)
objectRelatedToZoneNamePropertyValues = GetBuiltInPropertyIds ('IdAndCategories_RelatedZoneName', objects)
objectNamePropertyValues = GetBuiltInPropertyIds ('IdAndCategories_Name', objects)



objectLayerNames = GetPropertyValues(objectLayerPropertyValues)
#print('Error with Objects - Home Story Number:')
objectHomeStoryNumbers = GetPropertyValues(objectHomeStoryNumberPropertyValues)
objectRelatedToZoneNumbers = GetPropertyValues(objectRelatedToZoneNumberPropertyValues)
objectRelatedToZoneNames = GetPropertyValues(objectRelatedToZoneNamePropertyValues)
objectNames = GetPropertyValues(objectNamePropertyValues)

testLength = len(objectLayerNames)

if(testLength != len(objectHomeStoryNumbers)):
    print('Error with length of objectHomeStoryNumbers')
if(testLength != len(objectRelatedToZoneNumbers)):
    print('Error with length of objectRelatedToZoneNumbers')
if(testLength != len(objectRelatedToZoneNames)):
    print('Error with length of objectRelatedToZoneNames')
if(testLength != len(objectNames)):
    print('Error with length of objectNames')

    
'''
for i, prop in enumerate(objectHomeStoryNumberPropertyValues):
    if('normal' in objectHomeStoryNumberPropertyValues[i].propertyValues[0].propertyValue.status):
        print(str(objectNames[i]) + ' - ' + str(objectHomeStoryNumberPropertyValues[i].propertyValues[0].propertyValue.value))
        print(str(objectNames[i]) + ' - ' + str(objectHomeStoryNumbers[i]))
        print(str(objectNames[i]) + ' - ' + str(objectHomeStoryNumberPropertyValues[i].propertyValues[0].propertyValue.status))
    else:
        print('Error')
        print(str(objectNames[i]) + ' - ' + str(objectHomeStoryNumberPropertyValues[i].propertyValues[0].propertyValue.status))
'''
#filter our objects that are bathtubs, are part of furniture layer and are above Ground Floor
bathtubs = []
bathtub_RelatedToZoneNumbers = []
bathtub_RelatedToZoneNames = []

for i, currentObject in enumerate(objects):
    
    #objectLayerPropertyValue = GetBuiltInPropertyIds ('ModelView_LayerName', objects[i])
    currentLayerName = objectLayerNames[i]
    currentObjectName =objectNames[i]
    #if('normal' in objectHomeStoryNumberPropertyValues[i].propertyValues[0].propertyValue.status):
        #for some reason some objects cannot access Home Story Number in AC, we need to skip these
    if('Error' not in str(objectHomeStoryNumbers[i])):
        currentHomeStoryNumber = int(objectHomeStoryNumbers[i])
        for possibleLayerName in bathtub_PossibleLayerNames:
            if(currentLayerName == possibleLayerName):
                if(currentHomeStoryNumber >= 0):
                    if("Badewanne" in currentObjectName):
                        bathtubs.append(currentObject)
                        bathtub_RelatedToZoneNumbers.append(objectRelatedToZoneNumbers[i])
                        bathtub_RelatedToZoneNames.append(objectRelatedToZoneNames[i])
                        if('Wannenbad' not in objectRelatedToZoneNamePropertyValues[i].propertyValues[0].propertyValue.value ):
                            print('Following Bathtub is not in a correctly named Zone - it should be called Wannenbad:')
                            if(objectRelatedToZoneNumbers[i] == ''):
                                    print('No ZoneNumber has been found related to Bathtub')
                            else:
                                    print(objectRelatedToZoneNumbers[i])
                            print('Current name is: ' + str(objectRelatedToZoneNames[i]))
    else:
        if("Badewanne" in currentObjectName):
            print("Object reading error")
            print("Following object has been included, even though Home Story could not be read :") # the face that home story cannot be read has to do with Hotlinks. Reading all Bathtub objects without this might not be a problem. 
            print(objectNames[i])
            print(objectRelatedToZoneNumbers[i])
            print(objectHomeStoryNumbers[i])
            print('')
            bathtubs.append(currentObject)
            bathtub_RelatedToZoneNumbers.append(objectRelatedToZoneNumbers[i])
            bathtub_RelatedToZoneNames.append(objectRelatedToZoneNames[i])
            if('Wannenbad' not in objectRelatedToZoneNamePropertyValues[i].propertyValues[0].propertyValue.value ):
                print('Following Bathtub is not in a correctly named Zone - it should be called Wannenbad:')
                if(objectRelatedToZoneNumbers[i] == ''):
                        print('No ZoneNumber has been found related to Bathtub')
                else:
                        print(objectRelatedToZoneNumbers[i])
                print('Current name is: ' + str(objectRelatedToZoneNames[i]))





###################################### Start Data Processing #######################################

#create list of bools containing whether a zone is a "room" or not
apartmentRoomBools = []


#check whether zones are "rooms"

def CheckIfZonesAreRooms(zoneNameToBeChecked, possibleRoomNames):
    finalResult = 0
    for possibleRoomName in possibleRoomNames:
        if(possibleRoomName in zoneNameToBeChecked):
            finalResult = 1
     
    return finalResult
"""
#k=0
#for j, zoneName in enumerate(zoneNames):
#while k < len(zoneNames):
    # apartmentRoomBools.append(0)
   # l = 0
   # while l <  len(possibleRoomNames):
        #if(possibleRoomNames[l] in zoneNames[k]):
    #apartmentRoomBools.append( CheckIfZonesAreRooms(zoneName, possibleRoomNames))

        #l+=1
    #k+=1
"""
for  zoneName in apartmentZoneNames:
    apartmentRoomBools.append(CheckIfZonesAreRooms(zoneName, possibleRoomNames))




#create list with unique Zone Numbers
allUniqueApartmentNumbers = list(set(apartmentZoneNumbers))
allUniqueApartmentNumbers.sort()

'''
#leave only apartment numbers (delete Gewerbe/Abstellraume etc)

newUniqueList = []
for currentNumber in allUniqueApartmentNumbers:
    #l=0
   # append = False
    if len(currentNumber)>0 :
        for startingNumber in possibleApartmentStartingNumbers:
       # while l < len(possibleApartmentStartingNumbers):
            #print (k)
            #currentApartmentNumber = currentNumber
            firstCharacter = currentNumber[0]
            #print (firstCharacter)
            if str(startingNumber) in firstCharacter:
                #append = True
                newUniqueList.append(currentNumber)
                break
            #l+=1000
    
    #if(append):
    #    newUniqueList.append(currentNumber)
        

allUniqueApartmentNumbers = newUniqueList

'''

"""
#create list with zeroes for each Unique Apartment
apartmentRoomQuantity = []
#k=0

for k, x in enumerate(allUniqueApartmentNumbers):
#while k < len (allUniqueApartmentNumbers):
    apartmentRoomQuantity.append(0)
   # k+=1

"""

#Create empty dictionary with Apartment Numbers and Quantities of Rooms (for now only 0)

apartmentRoomAmountDict = dict.fromkeys(allUniqueApartmentNumbers, 0 )

#Create empty dictionary with Apartment Numbers and Quantities of Bathtubs (for now only 0)

apartmentBathtubAmountDict = dict.fromkeys(allUniqueApartmentNumbers, 0 )

#Create empty dictionary with Apartment Numbers and Area (for now only 0)

apartmentSizeDict = dict.fromkeys(allUniqueApartmentNumbers, 0 )

#Create empty dictionary with Apartment Numbers and Area (for now only 0)

apartmentWohnungsartDict = dict.fromkeys(allUniqueApartmentNumbers, '')

#Create empty dictionary with Apartment Numbers and Area (for now only 0)

apartmentBuildingNumberDict = dict.fromkeys(allUniqueApartmentNumbers, '')



#check and sum room quantities for apartments
for currentNumber in allUniqueApartmentNumbers:
    l=0
    while l < len (apartmentRoomBools):
        if currentNumber == apartmentZoneNumbers[l]:
            currentQuantityOfRooms = apartmentRoomAmountDict[currentNumber]
            apartmentRoomAmountDict[currentNumber] = currentQuantityOfRooms +  apartmentRoomBools[l]
        
        l+=1

#find groups without any "Rooms"
#print('groups without any "Rooms"')


#k=0
for currentKey  in list(apartmentRoomAmountDict.keys()):
    for currentValue in list(apartmentRoomAmountDict.values()):
        #myvalues = list( apartmentRoomAmountDict.values())
       # mykeys = list(apartmentRoomAmountDict.keys())
        if(currentValue == 0):
            print (currentKey)
            break
        #k+=1

#check if apartment has bathtub
for currentApartmentNumber in allUniqueApartmentNumbers:
    for currentObjectNumber in bathtub_RelatedToZoneNumbers:
        if currentObjectNumber == currentApartmentNumber:
            apartmentBathtubAmountDict[currentApartmentNumber] = 1


#calculate apartment area
for currentApartmentNumber in allUniqueApartmentNumbers:
    for i, currentZoneNumber in enumerate (apartmentZoneNumbers):
            if currentZoneNumber == currentApartmentNumber:
                currentSizeOfApartment = apartmentSizeDict[currentZoneNumber]
                apartmentSizeDict[currentZoneNumber] = currentSizeOfApartment + apartmentZoneAreas[i]

propID = acu.GetUserDefinedPropertyId("räume", "Wohnungstyp")
propertyValuesForZones = acc.GetPropertyValuesOfElements(apartmentZones, [propID])

detailedValue = acc.GetDetailsOfProperties([propID])


enumValues = []
i=0
j=1
for element in detailedValue[0].propertyDefinition.possibleEnumValues:
    if('Zimmer' in detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.displayValue):
        if(str(j) in detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.displayValue):
            enumValues.append(detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.enumValueId)
            j+=1
    i+=1
#print(enumValues)



#################################### Start Data Exporting #######################################
print('')
if overwriteProperties:
    i=0
    elemPropertyValues = []
    for zone in zones:
            j=0
            while j < len(list (apartmentRoomAmountDict.keys())):
                if apartmentZoneNumbers[i] == list (apartmentRoomAmountDict.keys())[j]:
                    x = list (apartmentRoomAmountDict.values())[j] - 1
                    propValue = act.NormalSingleEnumPropertyValue(enumValues[x]) # this needs to be checked, if property input list in Archicad changes
                    elemPropertyValues.append(act.ElementPropertyValue(
                        zone.elementId, propID, propValue))
                j+=1
            i+=1
    acc.SetPropertyValuesOfElements(elemPropertyValues)
else: 
    i=0
    print('Following Apartments need to be updated in Archicad with Quantity of Rooms: ')
    for zone in apartmentZones:
            j=0
            while j < len(list (apartmentRoomAmountDict.keys())):
                if apartmentZoneNumbers[i] == list (apartmentRoomAmountDict.keys())[j]:
                    x = list (apartmentRoomAmountDict.values())[j] - 1 # this only makes sense when '1-Zimmer' is in position [0] on the list
                    if ((propertyValuesForZones[i].propertyValues[0].propertyValue.value.displayValue) != enumValues[x].displayValue): # the property in Archicad is not the same as counted by the algorithm
                        print(str(list (apartmentRoomAmountDict.keys())[j]) + ' - '+str ( list (apartmentRoomAmountDict.values())[j]) + ' Zimmer')
                        print("At the moment it is "+ propertyValuesForZones[i].propertyValues[0].propertyValue.value.displayValue)
                        print('')
                        
                j+=1
            i+=1
    print('')

propID = acu.GetUserDefinedPropertyId("räume", "Wohnungsart")
propertyValuesForZones = acc.GetPropertyValuesOfElements(zones, [propID])

detailedValue = acc.GetDetailsOfProperties([propID])


wohnungsartPropList = ['gefördert', 'freifinanziert', 'Eigentum']
wohnungsartPropDict = dict.fromkeys(wohnungsartPropList, 0)

for i, element in enumerate (detailedValue[0].propertyDefinition.possibleEnumValues):
    if('gefördert' in detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.displayValue):
         wohnungsartPropDict["gefördert"] = detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.enumValueId
    if('freifinanziert' in detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.displayValue):
         wohnungsartPropDict["freifinanziert"] = detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.enumValueId
    if('eigentum' in detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.displayValue):
         wohnungsartPropDict["Eigentum"] = detailedValue[0].propertyDefinition.possibleEnumValues[i].enumValue.enumValueId
    

#ask if property should be overwritten - done
#overwrite Wohnungsart property - done
#print list if it should not be overwritten - done

#organize Wohnungsart into Dictionary
for currentApartmentNumber in allUniqueApartmentNumbers:
    for i, currentZoneNumber in enumerate(apartmentZoneNumbers):
        if(currentApartmentNumber == currentZoneNumber):
            apartmentWohnungsartDict[currentApartmentNumber] = apartmentZoneWohnungsart[i] # keep in mind that not all zones might have the same property value
            break


#check if apartment if foerderfahig                     #check this part, there might be errors!
elemPropertyValues = []
for currentApartmentNumber in allUniqueApartmentNumbers:
    if CheckApartmentSize(apartmentSizeDict[currentApartmentNumber], apartmentRoomAmountDict[currentApartmentNumber], apartmentBathtubAmountDict [currentApartmentNumber], bufferForApartmentArea): #if förderfähig then
        
        
            if overwriteProperties:
                
                apartmentWohnungsartDict[currentApartmentNumber] = "gefördert"

                for i, currentZone in enumerate(apartmentZones):
                    if(apartmentZoneNumbers[i] == currentApartmentNumber):

                        propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict["gefördert"])

                        elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
            #else:
                #apartmentWohnungsartDict[currentApartmentNumber] = apartmentZoneWohnungsart[i]
                #if('gefördert' in apartmentWohnungsartDict[currentApartmentNumber]):
                    #print('Following apartment förderfähig:')

    else:
        if overwriteProperties:
            for i, currentZone in enumerate(apartmentZones):

                if(apartmentZoneNumbers[i] == currentApartmentNumber):
                    if("Eigentum" not in apartmentZoneWohnungsart[i].propertyValues[0].propertyValue.value.displayValue): #if not eigentum then overwrite property to freifinanziert
                        propValue = act.NormalSingleEnumPropertyValue(wohnungsartPropDict['freifinanziert'])

                        apartmentWohnungsartDict[currentApartmentNumber] = 'freifinanziert'
                    else:
                        apartmentWohnungsartDict[currentApartmentNumber] = 'Eigentum'
                        

                        elemPropertyValues.append (act.ElementPropertyValue (currentZone.elementId, propID, propValue) )
        else:
            print('Following apartment not förderfähig:')
            print(currentApartmentNumber)
            print(str(apartmentRoomAmountDict[currentApartmentNumber]) + ' Zimmer - ' + str(apartmentSizeDict[currentApartmentNumber]) + ' m3')

            if(apartmentWohnungsartDict[currentApartmentNumber] != 'freifinanziert' and apartmentWohnungsartDict[currentApartmentNumber] != 'Eigentum'):
               print('Property needs to be corrected in Archicad!')

            else:
                print('Property is correctly set in Archicad')
            print('Current Property in Archicad -' + str(apartmentWohnungsartDict[currentApartmentNumber]))
            print('')
                
            

 
if overwriteProperties:
    acc.SetPropertyValuesOfElements(elemPropertyValues)



for currentZoneNumber in allUniqueApartmentNumbers:
    for i, currentZone in enumerate(apartmentZoneNumbers):
        if currentZoneNumber == currentZone:
            if apartmentBuildingNumberDict[currentZoneNumber] == '':
                apartmentBuildingNumberDict[currentZoneNumber] = apartmentBuildingNumber[i]




    


    #count apartmenttypes (number of rooms, foerderfahigkeit)
    
totalAmount_1RoomApt = 0
totalAmount_2RoomApt = 0
totalAmount_3RoomApt = 0
totalAmount_4RoomApt = 0
totalAmount_5RoomApt = 0

totalAmount_gefoerdert = 0
totalAmount_freifinanziert = 0
totalAmount_Eigentum = 0



totalAmount_gefoerdert_G1 = 0
totalAmount_freifinanziert_G1 = 0
totalAmount_Eigentum_G1 = 0

totalAmount_gefoerdert_G2 = 0
totalAmount_freifinanziert_G2 = 0
totalAmount_Eigentum_G2 = 0

totalAmount_gefoerdert_G3 = 0
totalAmount_freifinanziert_G3 = 0
totalAmount_Eigentum_G3 = 0

totalAmount_gefoerdert_G4 = 0
totalAmount_freifinanziert_G4 = 0
totalAmount_Eigentum_G4 = 0

totalAmount_gefoerdert_G5 = 0
totalAmount_freifinanziert_G5 = 0
totalAmount_Eigentum_G5 = 0

totalAmount_gefoerdert_G6 = 0
totalAmount_freifinanziert_G6 = 0
totalAmount_Eigentum_G6 = 0




totalArea_gefoerdert = 0
totalArea_freifinanziert = 0
totalArea_Eigentum = 0

totalArea_gefoerdert_G1 = 0
totalArea_freifinanziert_G1 = 0
totalArea_Eigentum_G1 = 0

totalArea_gefoerdert_G2 = 0
totalArea_freifinanziert_G2 = 0
totalArea_Eigentum_G2 = 0

totalArea_gefoerdert_G3 = 0
totalArea_freifinanziert_G3 = 0
totalArea_Eigentum_G3 = 0

totalArea_gefoerdert_G4 = 0
totalArea_freifinanziert_G4 = 0
totalArea_Eigentum_G4 = 0

totalArea_gefoerdert_G5 = 0
totalArea_freifinanziert_G5 = 0
totalArea_Eigentum_G5 = 0

totalArea_gefoerdert_G6 = 0
totalArea_freifinanziert_G6 = 0
totalArea_Eigentum_G6 = 0

for currentApartmentNumber in allUniqueApartmentNumbers:
    if(apartmentRoomAmountDict[currentApartmentNumber] == 1):
        totalAmount_1RoomApt += 1
    if(apartmentRoomAmountDict[currentApartmentNumber] == 2):
        totalAmount_2RoomApt += 1
    if(apartmentRoomAmountDict[currentApartmentNumber] == 3):
        totalAmount_3RoomApt += 1
    if(apartmentRoomAmountDict[currentApartmentNumber] == 4):
        totalAmount_4RoomApt += 1
    if(apartmentRoomAmountDict[currentApartmentNumber] == 5):
        totalAmount_5RoomApt += 1

    if("gefördert" in apartmentWohnungsartDict[currentApartmentNumber]):
        totalAmount_gefoerdert += 1
        totalArea_gefoerdert += apartmentSizeDict[currentApartmentNumber]

        if('1' in apartmentBuildingNumberDict[currentApartmentNumber]):
            #print('G1 - gefördert')
            totalAmount_gefoerdert_G1 += 1
            totalArea_gefoerdert_G1 += apartmentSizeDict[currentApartmentNumber]

        if('2' in apartmentBuildingNumberDict[currentApartmentNumber]):
            #print('G2 - gefördert')
            totalAmount_gefoerdert_G2 += 1
            totalArea_gefoerdert_G2 += apartmentSizeDict[currentApartmentNumber]

        if('3' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_gefoerdert_G3 += 1
            totalArea_gefoerdert_G3 += apartmentSizeDict[currentApartmentNumber]

        if('4' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_gefoerdert_G4 += 1
            totalArea_gefoerdert_G4 += apartmentSizeDict[currentApartmentNumber]

        if('5' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_gefoerdert_G5 += 1
            totalArea_gefoerdert_G5 += apartmentSizeDict[currentApartmentNumber]

        if('6' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_gefoerdert_G6 += 1
            totalArea_gefoerdert_G6 += apartmentSizeDict[currentApartmentNumber]


    if('freifinanziert' in apartmentWohnungsartDict[currentApartmentNumber]):
        totalAmount_freifinanziert += 1
        totalArea_freifinanziert += apartmentSizeDict[currentApartmentNumber]

        if('1' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_freifinanziert_G1 += 1
            totalArea_freifinanziert_G1 += apartmentSizeDict[currentApartmentNumber]

        if('2' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_freifinanziert_G2 += 1
            totalArea_freifinanziert_G2 += apartmentSizeDict[currentApartmentNumber]

        if('3' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_freifinanziert_G3 += 1
            totalArea_freifinanziert_G3 += apartmentSizeDict[currentApartmentNumber]

        if('4' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_freifinanziert_G4 += 1
            totalArea_freifinanziert_G4 += apartmentSizeDict[currentApartmentNumber]

        if('5' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_freifinanziert_G5 += 1
            totalArea_freifinanziert_G5 += apartmentSizeDict[currentApartmentNumber]

        if('6' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_freifinanziert_G6 += 1
            totalArea_freifinanziert_G6 += apartmentSizeDict[currentApartmentNumber]

    if('Eigentum' in apartmentWohnungsartDict[currentApartmentNumber]):
        totalAmount_Eigentum += 1
        totalArea_Eigentum += apartmentSizeDict[currentApartmentNumber]

        if('1' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_Eigentum_G1 += 1
            totalArea_Eigentum_G1 += apartmentSizeDict[currentApartmentNumber]
        if('2' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_Eigentum_G2 += 1
            totalArea_Eigentum_G2 += apartmentSizeDict[currentApartmentNumber]
        if('3' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_Eigentum_G3 += 1
            totalArea_Eigentum_G3 += apartmentSizeDict[currentApartmentNumber]
        if('4' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_Eigentum_G4 += 1
            totalArea_Eigentum_G4 += apartmentSizeDict[currentApartmentNumber]
        if('5' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_Eigentum_G5 += 1
            totalArea_Eigentum_G5 += apartmentSizeDict[currentApartmentNumber]
        if('6' in apartmentBuildingNumberDict[currentApartmentNumber]):
            totalAmount_Eigentum_G6 += 1
            totalArea_Eigentum_G6 += apartmentSizeDict[currentApartmentNumber]


#print(apartmentBuildingNumberDict)
#print(apartmentWohnungsartDict)

#export quantities to excel

if overwriteExcel:
    wb = load_workbook(filename = excelFileName)

    ws = wb["Stellplätze Fahrräder"]

    #C3 - G6 - Eigentum
    ws['C3'] = totalAmount_Eigentum_G6

    #C18 - G1-5 - gefördert + freifinanziert
    ws['C18'] = totalAmount_gefoerdert + totalAmount_freifinanziert # assuming there are none in G6



    ws2 = wb["Stellplätze PKW"]


    # C3 - G6 - Eigentum
    ws2['C3'] = totalAmount_Eigentum_G6

    # C11 - G1-5 - gefördert
    ws2['C11'] = totalAmount_gefoerdert

    # C18 - G1-5 - freifinanziert
    ws2['C18'] = totalAmount_freifinanziert

    ws3 = wb["Wohnflächen"]

    #Anzahl
    #D4 - G1 - gefördert
    ws3['D4'] = totalAmount_gefoerdert_G1
    #D5 - G1 - freifinanziert
    ws3['D5'] = totalAmount_freifinanziert_G1
    #D6 - G1 - Eigentum
    ws3['D6'] = totalAmount_Eigentum_G1
 
    #D11 - G2 - gefördert
    ws3['D11'] = totalAmount_gefoerdert_G2
    #D12 - G2 - freifinanziert
    ws3['D12'] = totalAmount_freifinanziert_G2
    #D13 - G2 - Eigentum
    ws3['D13'] = totalAmount_Eigentum_G2
 
    #D18 - G3 - gefördert
    ws3['D18'] = totalAmount_gefoerdert_G3
    #D19 - G3 - freifinanziert
    ws3['D19'] = totalAmount_freifinanziert_G3
    #D20 - G3 - Eigentum
    ws3['D20'] = totalAmount_Eigentum_G3
 
    #D25 - G4 - gefördert
    ws3['D25'] = totalAmount_gefoerdert_G4
    #D26 - G4 - freifinanziert
    ws3['D26'] = totalAmount_freifinanziert_G4
    #D27 - G4 - Eigentum
    ws3['D27'] = totalAmount_Eigentum_G4
 
    #D32 - G5 - gefördert
    ws3['D32'] = totalAmount_gefoerdert_G5
    #D33 - G5 - freifinanziert
    ws3['D33'] = totalAmount_freifinanziert_G5
    #D34 - G5 - Eigentum
    ws3['D34'] = totalAmount_Eigentum_G5
 
    #D39 - G6 - gefördert
    ws3['D39'] = totalAmount_gefoerdert_G6
    #D40 - G6 - freifinanziert
    ws3['D40'] = totalAmount_freifinanziert_G6
    #D41 - G6 - Eigentum
    ws3['D41'] = totalAmount_Eigentum_G6

    #Wohnfläche
    #C4 - G1 - gefördert
    ws3['C4'] = totalArea_gefoerdert_G1
    #C5 - G1 - freifinanziert
    ws3['C5'] = totalArea_freifinanziert_G1
    #C6 - G1 - Eigentum
    ws3['C6'] = totalArea_Eigentum_G1
 
    #C11 - G2 - gefördert
    ws3['C11'] = totalArea_gefoerdert_G2
    #C12 - G2 - freifinanziert
    ws3['C12'] = totalArea_freifinanziert_G2
    #C13 - G2 - Eigentum
    ws3['C13'] = totalArea_Eigentum_G2
 
    #C18 - G3 - gefördert
    ws3['C18'] = totalArea_gefoerdert_G3
    #C19 - G3 - freifinanziert
    ws3['C19'] = totalArea_freifinanziert_G3
    #C20 - G3 - Eigentum
    ws3['C20'] = totalArea_Eigentum_G3
 
    #C25 - G4 - gefördert
    ws3['C25'] = totalArea_gefoerdert_G4
    #C26 - G4 - freifinanziert
    ws3['C26'] = totalArea_freifinanziert_G4
    #C27 - G4 - Eigentum
    ws3['C27'] = totalArea_Eigentum_G4
 
    #C32 - G5 - gefördert
    ws3['C32'] = totalArea_gefoerdert_G5
    #C33 - G5 - freifinanziert
    ws3['C33'] = totalArea_freifinanziert_G5
    #C34 - G5 - Eigentum
    ws3['C34'] = totalArea_Eigentum_G5
 
    #C39 - G6 - gefördert
    ws3['C39'] = totalArea_gefoerdert_G6
    #C40 - G6 - freifinanziert
    ws3['C40'] = totalArea_freifinanziert_G6
    #C41 - G6 - Eigentum
    ws3['C41'] = totalArea_Eigentum_G6


    outputPath = os.path.join(scriptFolder, excelFileName)
    wb.save(outputPath)
